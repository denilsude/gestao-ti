<<<<<<< HEAD
=======
<<<<<<< HEAD
>>>>>>> e8a803335aee9a806ce299f6e50df1f1541b199d
import pandas as pd
import os
import tkinter as tk
from tkinter import PhotoImage, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from concurrent.futures import ThreadPoolExecutor
import webbrowser
import unicodedata
import re
import subprocess
import platform
import threading
import socket
import time
import csv



def limpar_tela():
    for widget in root.winfo_children():
        widget.destroy()

import tkinter as tk

def tela_inicial():
    limpar_tela()
    
    root.configure(bg="#003D1F")  # Fundo verde ainda mais escuro para maior contraste

       
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)
    
    # Cabeçalho
    header = tk.Frame(frame, bg="#026440", pady=10)  # Verdee intermediário para destaque
    header.pack(fill="x")
    
    label_titulo = tk.Label(header, text="Gestão de acesso Tecnologia - 3246", font=("Arial", 22, "bold"), bg="#026440", fg="white")
    label_titulo.pack(pady=10)
    
    # Área de Botões
    main_frame = tk.Frame(frame, bg="#003D1F")
    main_frame.pack(pady=20)
    
    botoes = [
        ("Colaboradores", cadastrar_ou_alterar_funcionario, "#008F68"),
        ("Acessar UEM", pesquisar_funcionario, "#00B386"),
        ("Checklist Admissão", checklist_admissao, "#00997B"),
        ("Checklist Demissão", checklist_desligamento, "#00997B"),
        ("Grupo de Acesso", grupo_acesso, "#008386"),
        ("Ip Scanner", varredura_ip, "#008386"),
        ("Relatórios", relatorio_geral, "#66CDAA")
    ]

    # Criar botões em 2 colunas
    for i, (text, command, color) in enumerate(botoes):
        row = i // 2   # Linha (vai mudando a cada 2 botões)
        col = i % 2    # Coluna (0 ou 1)
        
        tk.Button(
            main_frame,
            text=text,
            font=("Arial", 14, "bold"),
            width=25,
            height=2,
            bg=color,
            fg="white",
            relief="ridge",
            bd=3,
            command=command
        ).grid(row=row, column=col, padx=15, pady=15)

    # Centralizar as colunas
    main_frame.grid_columnconfigure(0, weight=1)
    main_frame.grid_columnconfigure(1, weight=1)
    
    # Rodapé
    footer = tk.Frame(frame, bg="#026440", pady=5)
    footer.pack(fill="x", side="bottom")
    
    label_footer = tk.Label(footer, text="Sicoob Credseguro - 2025", font=("Arial", 10), bg="#026440", fg="white")
    label_footer.pack()

def cadastrar_ou_alterar_funcionario():
    limpar_tela()

    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Cadastro/Alteração de Funcionário", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    tk.Label(frame, text="Digite nome, CPF, login ou e-mail para buscar:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(pady=10)
    entry_pesquisa = tk.Entry(frame, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    entry_pesquisa.pack(pady=10, ipady=5)
    entry_pesquisa.bind("<Return>", lambda event: buscar())

    form_frame = tk.Frame(frame, bg="#003D1F")
    form_frame.pack(pady=10)

    campos = ["CPF", "Nome", "Computador", "AD", "Login", "E-mail", "Consórcio", "Telefone", "IP","MAC" , "UEM"]
    entries = {}

    # Funções adicionais
    IP_FIXO = "10.201."

    DOMINIO_FIXO = ""

    def verificar_duplicidade(event=None):
        caminho_arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")  # ajuste se necessário

        if not os.path.exists(caminho_arquivo):
            return

        try:
            df = pd.read_excel(caminho_arquivo)

            # Normalizar colunas (garantir que existam)
            colunas_esperadas = ["Nome", "CPF", "IP"]
            for col in colunas_esperadas:
                if col not in df.columns:
                    return

        except Exception as e:
            messagebox.showerror("Erro ao verificar duplicidade", f"Erro ao ler o arquivo: {e}")
    def remover_acentos(texto):
        return ''.join(
            c for c in unicodedata.normalize('NFKD', texto)
            if not unicodedata.combining(c)
        )
    def preencher_ad(event=None):
        nome_completo = entries["Nome"].get().strip()

        if not nome_completo:
            return

        nome_sem_acentos = remover_acentos(nome_completo)
        nomes = re.findall(r'\b\w+\b', nome_sem_acentos.lower())  # pega só palavras com letras/números

        if len(nomes) >= 2:
            primeiro = nomes[0]
            ultimo = nomes[-1]
            ad_formatado = f"{primeiro}.{ultimo}"
        elif nomes:
            ad_formatado = nomes[0]
        else:
            ad_formatado = ""

        entries["AD"].delete(0, tk.END)
        entries["AD"].insert(0, ad_formatado)
    def aplicar_mascara_mac(event):
        entry = entries["MAC"]
        texto = entry.get().upper()
        texto_hex = ''.join(filter(lambda c: c in "0123456789ABCDEF", texto))[:12]

        formatado = ":".join(texto_hex[i:i+2] for i in range(0, len(texto_hex), 2))

        entry.delete(0, tk.END)
        entry.insert(0, formatado)
    def formatar_nome(event):
        entry = entries["Nome"]
        texto = entry.get()
        cursor_pos = entry.index(tk.INSERT)

        # Não faz nada se campo estiver vazio
        if not texto:
            return

        # Preserva espaços múltiplos entre palavras enquanto digita
        palavras = texto.split(' ')
        palavras_formatadas = [p.capitalize() if p else '' for p in palavras]
        texto_corrigido = ' '.join(palavras_formatadas)

        # Atualiza o campo só se houver mudança
        if texto != texto_corrigido:
           entry.delete(0, tk.END)
           entry.insert(0, texto_corrigido)
           entry.icursor(min(cursor_pos, len(texto_corrigido)))

    def limpar_nome(event):
        entry = entries["Nome"]
        texto = entry.get()
        texto_filtrado = ''.join(
            char for char in texto if char.isalpha() or char.isspace() or unicodedata.category(char).startswith('L')
        )
        entry.delete(0, tk.END)
        entry.insert(0, texto_filtrado)
    
    def formatar_maiusculo(event):
        entry = event.widget
        texto = entry.get()
        entry.delete(0, tk.END)
        entry.insert(0, texto.upper())
    def aplicar_mascara_telefone(event=None):
        entry = event.widget
        texto = entry.get()
        numeros = ''.join(filter(str.isdigit, texto))[:11]

        if len(numeros) <= 2:
            formatado = f"({numeros}"
        elif len(numeros) <= 7:
            formatado = f"({numeros[:2]}){numeros[2:]}"
        else:
            formatado = f"({numeros[:2]}){numeros[2:7]}-{numeros[7:]}"

        entry.delete(0, tk.END)
        entry.insert(0, formatado)
    
    def aplicar_mascara_cpf(event):
        texto = entries["CPF"].get()
        texto_numerico = ''.join(filter(str.isdigit, texto))[:11]
        novo_texto = ""

        for i in range(len(texto_numerico)):
            if i == 3 or i == 6:
                novo_texto += "."
            elif i == 9:
                novo_texto += "-"
            novo_texto += texto_numerico[i]

        entries["CPF"].delete(0, tk.END)
        entries["CPF"].insert(0, novo_texto)

    def cpf_valido(cpf):
        cpf = ''.join(filter(str.isdigit, cpf))
        if len(cpf) != 11 or cpf == cpf[0] * 11:
            return False
        soma1 = sum(int(cpf[i]) * (10 - i) for i in range(9))
        digito1 = (soma1 * 10 % 11) % 10
        soma2 = sum(int(cpf[i]) * (11 - i) for i in range(10))
        digito2 = (soma2 * 10 % 11) % 10
        return cpf[-2:] == f"{digito1}{digito2}"

    def validar_cpf(event):
        cpf_texto = entries["CPF"].get()
        if not cpf_valido(cpf_texto):
            messagebox.showerror("CPF Inválido", "Digite um CPF válido.")
            entries["CPF"].focus_set()

    
    def manter_sufixo_email(event):
        texto = entries["E-mail"].get()
        if DOMINIO_FIXO in texto:
            prefixo = texto.split(DOMINIO_FIXO)[0]
        else:
            prefixo = texto.replace(DOMINIO_FIXO, ". ")
        novo_email = prefixo + DOMINIO_FIXO
        entries["E-mail"].delete(0, tk.END)
        entries["E-mail"].insert(0, novo_email)
        entries["E-mail"].icursor(len(prefixo))

    def validar_email(event):
        email = entries["E-mail"].get()
        if not email or email == DOMINIO_FIXO:
            messagebox.showerror("E-mail inválido", "Não precisa digitar '@sicoob.com.br'")
            entries["E-mail"].focus_set()

    def manter_prefixo_ip(event):
        texto = entries["IP"].get()
        sufixo = texto.replace(IP_FIXO, "")
        numeros = ''.join(filter(str.isdigit, sufixo))[:6]

        bloco1 = numeros[:2]
        bloco2 = numeros[2:]

        novo_ip = IP_FIXO
        if bloco1:
            novo_ip += bloco1
        if bloco2:
            novo_ip += "." + bloco2

        entries["IP"].delete(0, tk.END)
        entries["IP"].insert(0, novo_ip)
        entries["IP"].icursor(tk.END)

    def validar_ip(event):
        ip = entries["IP"].get()
        partes = ip.split(".")
        if len(partes) != 4:
            messagebox.showerror("IP Inválido", "O IP deve ter o formato 10.201.xx.xxx")
            entries["IP"].focus_set()
            return
        try:
            for parte in partes:
                if not parte.isdigit() or not (0 <= int(parte) <= 255):
                    raise ValueError
        except:
            messagebox.showerror("IP Inválido", "Cada parte do IP deve estar entre 0 e 255.")
            entries["IP"].focus_set()
    
    def somente_letras(event):
        texto = entries["Nome"].get()

        # Permitir letras, acentos e espaços
        texto_filtrado = ''.join(char for char in texto if (
        char.isalpha() or char.isspace() or unicodedata.category(char).startswith('L')
    ))

        # Atualiza o campo com apenas caracteres permitidos
        if texto != texto_filtrado:
            entries["Nome"].delete(0, tk.END)
            entries["Nome"].insert(0, texto_filtrado)
    
    for i, campo in enumerate(campos):
        row, col = divmod(i, 2)
        tk.Label(
            form_frame, text=f"{campo}:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w"
        ).grid(row=row, column=col * 2, sticky="w", padx=10, pady=5)

        entry = tk.Entry(
            form_frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F"
        )
        entry.grid(row=row, column=col * 2 + 1, padx=10, pady=1, ipady=5, sticky="w")
        entries[campo] = entry

        if campo == "CPF":
            entry.bind("<KeyRelease>", aplicar_mascara_cpf)
            entry.bind("<FocusOut>", validar_cpf)
            entry.bind("<FocusOut>", verificar_duplicidade)

        if campo == "E-mail":
            #entry.insert(0, DOMINIO_FIXO)
            #entry.bind("<KeyRelease>", manter_sufixo_email)
            #entry.bind("<FocusOut>", validar_email)
            pass

        if campo == "IP":
            entry.insert(0, IP_FIXO)
            entry.bind("<KeyRelease>", manter_prefixo_ip)
            entry.bind("<FocusOut>", validar_ip)
            entry.bind("<FocusOut>", verificar_duplicidade)

        if campo == "Nome" :
            entry.bind("<KeyRelease>", formatar_nome)
            entry.bind("<FocusOut>", limpar_nome)
            entry.bind("<KeyRelease>", preencher_ad)
            entry.bind("<FocusOut>", verificar_duplicidade)
                    
        if campo == "Consórcio":
            entry.insert(0, "Não tem")

        if campo == "Login" :
            entry.insert(0, "")    

        if campo in ["Login", "Computador"]:
            entry.bind("<KeyRelease>", formatar_maiusculo)
            entry.bind("<FocusOut>", formatar_maiusculo)
            entry.bind("<FocusOut>", verificar_duplicidade)
        
        if campo == "Telefone":
            entry.insert(0, "(62)3275-0200")
            entry.bind("<KeyRelease>", aplicar_mascara_telefone)

        if campo == "MAC":
            entry.bind("<KeyRelease>", aplicar_mascara_mac)

                    
    # Base da linha onde os campos adicionais começam
    linha_base = (len(campos) + 1) // 2

    # PA
    pa_var = tk.StringVar()
    valores_pa = [" ", "UAD", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"]
    tk.Label(form_frame, text="PA:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base, column=0, sticky="w", padx=10, pady=5)

    pa_combobox = ttk.Combobox(form_frame, textvariable=pa_var, values=valores_pa, state="readonly", font=("Arial", 12), width=38)
    pa_combobox.grid(row=linha_base, column=1, padx=10, pady=5, ipady=2)
    entries["PA"] = pa_combobox

    # Prognum
    prognum_var = tk.StringVar()
    valores_prognum = [" ", "Sim", "Não"]
    tk.Label(form_frame, text="Prognum:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base, column=2, sticky="w", padx=10, pady=5)

    prognum_combobox = ttk.Combobox(form_frame, textvariable=prognum_var, values=valores_prognum, state="readonly", font=("Arial", 12), width=38)
    prognum_combobox.grid(row=linha_base, column=3, padx=10, pady=5, ipady=2)
    entries["Prognum"] = prognum_combobox

    # Cargo
    cargo_var = tk.StringVar()
    valores_cargo = [" ", "Auxiliar", "Assistente", "Agente de atendimento", "Agente de relacionamento", "Analista", "Caixa", "Conselheiro(a)", "Diretor", "Estagiário", "Especialista", "Gerente de relacionamento", "Gestor", "Jovem aprendiz", "Presidente", "Secretaria", "Superitendente", "Supervisor(a)", "Tesoureiro(a)"]
    tk.Label(form_frame, text="Cargo:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base + 1, column=0, sticky="w", padx=10, pady=5)
 
    cargo_combobox = ttk.Combobox(form_frame, textvariable=cargo_var, values=valores_cargo, state="readonly", font=("Arial", 12), width=38)
    cargo_combobox.grid(row=linha_base + 1, column=1, padx=10, pady=5, ipady=2)
    entries["Cargo"] = cargo_combobox

    # Setor
    setor_var = tk.StringVar()
    valores_setor = [" ", "Administrativo", "Advocacia", "Cadastro", "Crédito", "Cobrança", "Comercial", "Conselho","Controladoria", "Diretoria", "Financeiro", "Financeiro UAD", "Secretária", "Gente e Gestão", "Marketing", "Produtos", "Tecnologia","Superitendencia", "Sustentabilidade", "Riscos e Controles"]
    tk.Label(form_frame, text="Setor:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base + 1, column=2, sticky="w", padx=10, pady=5)

    setor_combobox = ttk.Combobox(form_frame, textvariable=setor_var, values=valores_setor, state="readonly", font=("Arial", 12), width=38)
    setor_combobox.grid(row=linha_base + 1, column=3, padx=10, pady=5, ipady=2)
    entries["Setor"] = setor_combobox

    # O metodo buscar reflete na tela de cadastro de novo colaborador
    def buscar():        
        termo = entry_pesquisa.get().strip().lower()
        if not termo:
            messagebox.showerror("Erro", "Digite um termo para pesquisa!")
            
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):

            df = pd.read_excel(arquivo, dtype=str).fillna("")
            resultado = df[df.apply(lambda row: any(termo in str(valor).lower() for valor in row), axis=1)]

            if not resultado.empty:
                dados = resultado.iloc[0]
               # Preenchendo os campos de texto normais
                for campo in campos:
                    entries[campo].delete(0, tk.END)
                    entries[campo].insert(0, dados[campo])
            # Preenchendo o campo "PA"
                if "PA" in dados and dados["PA"] in valores_pa:
                    pa_var.set(dados["PA"])
                else:
                    pa_var.set(" ")
            # Preenchendo o campo "Cargo"
                if "Cargo" in dados and dados["Cargo"] in valores_cargo:
                    cargo_var.set(dados["Cargo"])
                else:
                    cargo_var.set(" ")

            # Preenchendo o campo "Setor"
                if "Setor" in dados and dados["Setor"] in valores_setor:
                    setor_var.set(dados["Setor"])
                else:
                    setor_var.set(" ")
            # Preenchendo corretamente o campo "Prognum" na Combobox
                if "Prognum" in dados and dados["Prognum"] in ["Sim", "Não"]:
                    prognum_var.set(dados["Prognum"])  # Atualiza a Combobox com o valor correto
                else:
                    prognum_var.set("erro")  # Define "Sim" como padrão caso o valor não seja válido

            else:
                messagebox.showinfo("Pesquisa", "Nenhum funcionário encontrado.")
        else:
             messagebox.showerror("Erro", "Arquivo não encontrado.")

    def excluir_funcionario():
        cpf = entries["CPF"].get().strip()
        cpf_numerico = ''.join(filter(str.isdigit, cpf))

        if not cpf_numerico:
            messagebox.showerror("Erro", "Nenhum CPF informado. Faça a busca de um funcionário primeiro.")
            return

        confirmacao = messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir o funcionário com CPF {cpf}?")
        if not confirmacao:
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):
            df = pd.read_excel(arquivo, dtype=str).fillna("")
            df_filtrado = df[df["CPF"].str.replace(r'\D', '', regex=True) != cpf_numerico]

            if len(df) == len(df_filtrado):
                messagebox.showinfo("Não encontrado", "Funcionário não encontrado para exclusão.")
            else:
                df_filtrado.to_excel(arquivo, index=False)
                messagebox.showinfo("Excluído", "Funcionário excluído com sucesso.")
                tela_inicial()
        else:
            messagebox.showerror("Erro", "Arquivo de funcionários não encontrado.")
    def salvar_alteracoes():
        novo_dados = {campo: entries[campo].get() for campo in campos}
        campos_reordenados = ["CPF", "Nome", "Computador", "AD", "Login", "E-mail", "Prognum", "Consórcio", "PA", "Setor", "Telefone", "IP", "MAC", "Cargo", "UEM"]
        novo_dados["PA"] = pa_var.get()
        novo_dados["Prognum"] = prognum_var.get()
        novo_dados["Cargo"] = cargo_var.get()
        novo_dados["Setor"] = setor_var.get()         

        if any(not valor for valor in novo_dados.values()):
            messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):
            # Lê todas as abas
            all_sheets = pd.read_excel(arquivo, sheet_name=None, dtype=str)
            all_sheets = {nome: df.fillna("") for nome, df in all_sheets.items()}

            nome_aba_principal = "Ativos"  # substitua pelo nome real da aba principal, se necessário
            df = all_sheets.get(nome_aba_principal)

            if df is None:
                messagebox.showerror("Erro", f"A aba '{nome_aba_principal}' não foi encontrada no arquivo!")
                return

            if list(df.columns) != campos_reordenados:
                messagebox.showerror("Erro", "A ordem das colunas na planilha não está correta!")
                return 
            
            index = df[df["CPF"] == novo_dados["CPF"]].index
            if not index.empty:
                df.loc[index[0], :] = [novo_dados[campo] for campo in campos_reordenados]
                all_sheets[nome_aba_principal] = df
            else:
                 # Adiciona novo funcionário
                novo_df = pd.DataFrame([novo_dados], columns=campos_reordenados)
                all_sheets[nome_aba_principal] = pd.concat([df, novo_df], ignore_index=True)

            # Salva todas as abas de volta no Excel
            with pd.ExcelWriter(arquivo, engine='openpyxl', mode='w') as writer:
                for aba, data in all_sheets.items():
                    data.to_excel(writer, sheet_name=aba, index=False)

            messagebox.showinfo("Sucesso", "Alterações salvas com sucesso!")
            tela_inicial()
        else:
            messagebox.showerror("Erro", "Arquivo não encontrado!")

    # Frame para botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)
    
    botoes = [
       # ("Buscar", buscar, "#008F68"),
        ("Salvar", salvar_alteracoes, "#00B386"),
        ("Excluir", excluir_funcionario, "#00B386"),
        ("Voltar", tela_inicial, "#00997B")
    ]
    
    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)


def grupo_acesso():
    limpar_tela()

    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Grupo de acesso", font=("Arial", 18, "bold"),bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de pesquisa (nome, CPF, e-mail ou login)
    global entry_pesquisa
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite o nome", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa = tk.Entry(linha_pesquisa, font=("Arial", 12), width=26)
    entry_pesquisa.pack(side="left")
    entry_pesquisa.bind("<Return>", lambda event: buscar_funcionario_grupo())
   
    global entry_login
    entry_login = tk.Entry(frame, font=("Arial", 12), width=42, state="readonly")
    entry_login.pack(pady=5)

    # Carrega grupos do Excel
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    grupos = {}

    if os.path.exists(arquivo):
        try:
            df = pd.read_excel(arquivo, sheet_name="Grupo")
            for _, row in df.iterrows():
                cargo = str(row[0]).strip()
                grupo = str(row[1]).strip()
                if cargo and grupo:
                    grupos[cargo] = grupo
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler a planilha: {e}")
            return
    else:
        messagebox.showerror("Erro", "Arquivo 'funcionarios.xlsx' não encontrado.")
        return

    if not grupos:
        messagebox.showinfo("Aviso", "Nenhum grupo encontrada na aba 'Grupo'.")
        return

    # Dropdown de cargos
    combo_frame = tk.Frame(frame, bg="#003D1F")
    combo_frame.pack(pady=10)

    tk.Label(combo_frame, text="Escolha a cargo:", font=("Arial", 14, "bold"),  bg="#003D1F", fg="white").pack(side="left", padx=(0,10))

    combo = ttk.Combobox(combo_frame, values=list(grupos.keys()), state="readonly", font=("Arial", 12), width=20)
    combo.pack(side="left")

    # Caixa de texto com scrollbar (tamanho reduzido)
    text_frame = tk.Frame(frame, bg="#003D1F")
    text_frame.pack(pady=10)

    caixa_texto = tk.Text(text_frame, wrap="word", yscrollcommand=lambda *args: scrollbar.set(*args),
                      font=("Arial", 12), bg="#E8F5E9", fg="#003D1F", relief="ridge", bd=3,
                      width=90, height=22)
    caixa_texto.pack(side="left")

    scrollbar = tk.Scrollbar(text_frame, command=caixa_texto.yview)
    scrollbar.pack(side="right", fill="y")

    def mostrar_grupo(event=None):
        cargo = combo.get()
        grupo = grupos.get(cargo, "grupo não encontrado.")
        caixa_texto.delete("1.0", tk.END)
        caixa_texto.insert(tk.END, grupo)

    combo.bind("<<ComboboxSelected>>", mostrar_grupo)

    # Botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    def copiar_texto():
        texto = caixa_texto.get("1.0", tk.END).strip()
        if texto:
            root.clipboard_clear()
            root.clipboard_append(texto)
            messagebox.showinfo("Copiado", "Texto copiado para a área de transferência.")

    def salvar_texto():
        cargo = combo.get()
        novo_grupo = caixa_texto.get("1.0", tk.END).strip()
        if not cargo:
            messagebox.showwarning("Aviso", "Selecione uma cargo antes de salvar.")
            return
        try:
            df = pd.read_excel(arquivo, sheet_name="Grupo")
            cargo_encontrada = False
            for i in range(len(df)):
                if str(df.iloc[i, 0]).strip() == cargo:
                    df.iat[i, 1] = novo_grupo
                    cargo_encontrada = True
                    break
            if cargo_encontrada:
                with pd.ExcelWriter(arquivo, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df.to_excel(writer, sheet_name="Grupo", index=False)
                messagebox.showinfo("Sucesso", "Grupo atualizado com sucesso.")
            else:
                messagebox.showwarning("Aviso", "Cargo não encontrado na planilha.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar: {e}")

    botoes = [
        ("Copiar", copiar_texto, "#026440"),
        ("Salvar", salvar_texto, "#008F68"),
        ("Voltar", tela_inicial, "#00997B")
    ]

    for i, (texto, comando, cor) in enumerate(botoes):
        tk.Button(botoes_frame, text=texto, font=("Arial", 12, "bold"), width=20, height=2,
                  bg=cor, fg="white", relief="ridge", bd=3, command=comando).grid(row=0, column=i, padx=10)
def pesquisar_funcionario():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Acesso Remoto UEM", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    tk.Label(frame, text="Digite nome do usuário:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(pady=10)
    entry_pesquisa = tk.Entry(frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    entry_pesquisa.pack(padx=20, pady=10, ipady=5)

    # Associa a tecla "Enter" à função de busca
    entry_pesquisa.bind("<Return>", lambda event: buscar_uem())

    # Adicionando Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side="right", fill="y")

    colunas = ["Computador", "IP", "Nome", "Setor", "PA", "MAC", "Cargo", "UEM"]
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=10, yscrollcommand=tree_scroll.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=200, anchor="center")

    tree_scroll.config(command=tree.yview)
    tree.pack(expand=True, fill="both")

    def abrir_link(event):
        item = tree.selection()
        if item:
            valores = tree.item(item, "values")
            link = valores[-1]  # Última coluna (UEM)
            if link.startswith("http"):
                webbrowser.open(link)

    tree.bind("<Double-1>", abrir_link)

    def buscar_uem():
        for item in tree.get_children():
            tree.delete(item)

        uem = entry_pesquisa.get().strip().lower()
        if not uem:
            messagebox.showerror("Erro", "Digite um termo para pesquisa!")
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):
            try:
                # Ler as abas "Ativos" e "PC"
                df_ativos = pd.read_excel(arquivo, sheet_name="Ativos", dtype=str).fillna("")
                df_pc = pd.read_excel(arquivo, sheet_name="PC", dtype=str).fillna("")

                # Filtrar resultados nas duas abas
                resultados_ativos = df_ativos[df_ativos.apply(lambda row: any(uem in str(valor).lower() for valor in row), axis=1)]
                resultados_pc = df_pc[df_pc.apply(lambda row: any(uem in str(valor).lower() for valor in row), axis=1)]

                # Concatenar os resultados
                resultados = pd.concat([resultados_ativos, resultados_pc], ignore_index=True)
                
                if not resultados.empty:
                    for _, row in resultados.iterrows():
                        tree.insert("", "end", values=[row[col] for col in tree['columns']])
                else:
                    messagebox.showinfo("Pesquisa", "Nenhum funcionário encontrado.")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        else:
            messagebox.showerror("Erro", "Arquivo não encontrado.")

    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    botoes = [
        ("Buscar", buscar_uem, "#008F68"),
        ("Voltar", tela_inicial, "#00997B")
    ]

    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)

def ping_ip(ip):
    param = "-n" if platform.system().lower() == "windows" else "-c"
    comando = ["ping", param, "1", ip]
    try:
        subprocess.check_output(comando, stderr=subprocess.DEVNULL, universal_newlines=True, timeout=1)
        return True
    except subprocess.CalledProcessError:
        return False
    except subprocess.TimeoutExpired:
        return False

def obter_hostname(ip):
    try:
        return socket.gethostbyaddr(ip)[0]
    except socket.herror:
        return ""
    
def varredura_ip():
    limpar_tela()

    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Varredura de IP na Rede", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo faixa IP
    instrucao = tk.Label(frame, text="Informe a faixa de IP (ex: 10.201.60.1-10.201.60.254):", font=("Arial", 12), bg="#003D1F", fg="white")
    instrucao.pack(pady=5)
    entry_faixa = tk.Entry(frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9")
    entry_faixa.pack(pady=5, ipady=3)

    # Campo filtro hostname
    filtro_label = tk.Label(frame, text="Filtrar por Hostname (opcional):", font=("Arial", 12), bg="#003D1F", fg="white")
    filtro_label.pack(pady=5)
    entry_filtro = tk.Entry(frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9")
    entry_filtro.pack(pady=5, ipady=3)

    # Barra de progresso
    progress_var = tk.DoubleVar()
    progress = ttk.Progressbar(frame, variable=progress_var, maximum=100)
    progress.pack(fill="x", pady=10)

    # Tabela Treeview
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(fill="both", expand=True, pady=10)

    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side="right", fill="y")

    tree = ttk.Treeview(tree_frame, columns=("IP", "Host"), show="headings", yscrollcommand=tree_scroll.set)
    tree.heading("IP", text="Endereço IP")
    tree.heading("Host", text="Hostname")
    tree.column("IP", width=150)
    tree.column("Host", width=300)
    tree.pack(fill="both", expand=True)

    # Logs detalhados
    log_frame = tk.Frame(frame)
    log_frame.pack(fill="both", expand=False, pady=5)
    
    text_log_scroll = tk.Scrollbar(log_frame)
    text_log_scroll.pack(side="right", fill="y")

    text_log = tk.Text(log_frame, height=10, bg="black", fg="lime", font=("Courier", 10), yscrollcommand=text_log_scroll.set)
    text_log.pack(fill="both", expand=True)

    text_log_scroll.config(command=text_log.yview)

    btn_frame = tk.Frame(frame, bg="#003D1F")
    btn_frame.pack(pady=10)

    resultado = []

    def log(msg):
        text_log.insert("end", msg + "\n")
        text_log.see("end")  # Scroll automático

    def iniciar_varredura():
        faixa = entry_faixa.get().strip()
        filtro = entry_filtro.get().strip().lower()
        for i in tree.get_children():
            tree.delete(i)
        resultado.clear()
        progress_var.set(0)
        text_log.delete("1.0", "end")

        if "-" not in faixa:
            messagebox.showerror("Erro", "Informe a faixa corretamente (ex: 192.168.1.1-192.168.1.254).")
            return

        ip_inicio, ip_fim = faixa.split("-")
        partes_inicio = ip_inicio.strip().split(".")
        partes_fim = ip_fim.strip().split(".")
        if partes_inicio[:-1] != partes_fim[:-1]:
            messagebox.showerror("Erro", "O intervalo deve estar na mesma sub-rede.")
            return

        prefixo = ".".join(partes_inicio[:-1]) + "."
        ini = int(partes_inicio[-1])
        fim = int(partes_fim[-1])
        total = fim - ini + 1

        encontrados = []

        def escanear_ip(i):
            ip = prefixo + str(i)
            log(f"Pinging {ip}...")
            if ping_ip(ip):
                host = obter_hostname(ip)
                log(f"✔ {ip} respondeu - Hostname: {host}")
                if filtro == "" or filtro in host.lower():
                    encontrados.append((ip, host))
                    tree.insert("", "end", values=(ip, host))
            else:
                log(f"✖ {ip} não respondeu")
            progress_var.set((i - ini + 1) / total * 100)

        def escanear():
            btn_iniciar.config(state="disabled")
            btn_exportar.config(state="disabled")
            with ThreadPoolExecutor(max_workers=50) as executor:
                futures = [executor.submit(escanear_ip, i) for i in range(ini, fim + 1)]
                for f in futures:
                    f.result()
            resultado.extend(encontrados)
            log("✅ Varredura finalizada!")
            progress_var.set(100)
            btn_iniciar.config(state="normal")
            btn_exportar.config(state="normal")

        threading.Thread(target=escanear).start()

    def exportar_csv():
        if not resultado:
            messagebox.showinfo("Informação", "Nenhum dado para exportar.")
            return
        nome_arquivo = f"varredura_{time.strftime('%Y%m%d_%H%M%S')}.csv"
        with open(nome_arquivo, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Endereço IP", "Hostname"])
            writer.writerows(resultado)
        messagebox.showinfo("Exportado", f"Arquivo salvo: {nome_arquivo}")

    btn_iniciar = tk.Button(btn_frame, text="Iniciar Varredura", font=("Arial", 12, "bold"), bg="#00B386", fg="white", command=iniciar_varredura)
    btn_iniciar.pack(side="left", padx=5)

    btn_exportar = tk.Button(btn_frame, text="Exportar CSV", font=("Arial", 12, "bold"), bg="#00997B", fg="white", command=exportar_csv, state="disabled")
    btn_exportar.pack(side="left", padx=5)

    btn_voltar = tk.Button(btn_frame, text="Voltar", font=("Arial", 12, "bold"), bg="#008F68", fg="white", command=tela_inicial)
    btn_voltar.pack(side="left", padx=5)

def relatorio_geral():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Relatório Geral de Funcionários", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de busca
    filtro_var = tk.StringVar()
    filtro_entry = tk.Entry(frame, textvariable=filtro_var, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    filtro_entry.pack(pady=10, ipady=5)
    
    
    def aplicar_filtro():
        query = filtro_var.get().lower()
        tree.delete(*tree.get_children())
        for _, row in df.iterrows():
            if any(query in str(row[col]).lower() for col in tree['columns']):
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
    
    tk.Button(frame, text="Filtrar", font=("Arial", 12, "bold"), width=20, height=2, bg="#008F68", fg="white", relief="ridge", bd=3, command=aplicar_filtro).pack(pady=10)

    
    # Criando um Frame para conter a Treeview e a Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    # Adicionando Scrollbars horizontal e vertical
    tree_scroll_y = tk.Scrollbar(tree_frame, orient="vertical")
    tree_scroll_y.pack(side="right", fill="y")

    tree_scroll_x = tk.Scrollbar(tree_frame, orient="horizontal")
    tree_scroll_x.pack(side="bottom", fill="x")

    colunas = ["CPF", "Nome", "Computador", "AD", "Login", "E-mail", "Prognum", "Consórcio", "PA", "Setor", "Cargo", "Telefone", "IP", "MAC"]
    
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=15, 
                        yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=120, anchor="center")

    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    tree.pack(expand=True, fill="both")
   
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    global df
    if os.path.exists(arquivo):
        df = pd.read_excel(arquivo, dtype=str).fillna("")
        
        if not df.empty:
            for _, row in df.iterrows():
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
        else:
            messagebox.showinfo("Relatório", "Nenhum funcionário cadastrado.")
    else:
        messagebox.showerror("Erro", "Arquivo não encontrado.")

    tk.Button(frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=tela_inicial).pack(side="right", padx=10, pady=10)
def relatorio_admissao():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Relatório Geral de Admitidos", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de busca
    filtro_var = tk.StringVar()
    filtro_entry = tk.Entry(frame, textvariable=filtro_var, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    filtro_entry.pack(pady=10, ipady=5)
    
    
    def aplicar_filtro():
        query = filtro_var.get().lower()
        tree.delete(*tree.get_children())
        for _, row in df.iterrows():
            if any(query in str(row[col]).lower() for col in tree['columns']):
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
    
    tk.Button(frame, text="Filtrar", font=("Arial", 12, "bold"), width=20, height=2, bg="#008F68", fg="white", relief="ridge", bd=3, command=aplicar_filtro).pack(pady=10)

    
    # Criando um Frame para conter a Treeview e a Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    # Adicionando Scrollbars horizontal e vertical
    tree_scroll_y = tk.Scrollbar(tree_frame, orient="vertical")
    tree_scroll_y.pack(side="right", fill="y")

    tree_scroll_x = tk.Scrollbar(tree_frame, orient="horizontal")
    tree_scroll_x.pack(side="bottom", fill="x")

    colunas = ["Funcionário","Criar E-mail","Criar AD","Criar SISBR","Criar SIPAG","Criar ASSINATURA","Ramal WEBEX","Portal CCS A/O","CX compartilhada/lista dpt","Colocar lista3246.todos","Acessos Sisbr para central","Grupos outlook","Termo notebook","Cadastrar biometria"]
    
       
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=15, 
                        yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=120, anchor="center")

    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    tree.pack(expand=True, fill="both")
   
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    global df
    if os.path.exists(arquivo):
        df = pd.read_excel(arquivo, sheet_name="Admissao", dtype=str).fillna("")
        
        if not df.empty:
            for _, row in df.iterrows():
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
        else:
            messagebox.showinfo("Relatório", "Nenhum funcionário cadastrado.")
    else:
        messagebox.showerror("Erro", "Arquivo não encontrado.")

    tk.Button(frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=tela_inicial).pack(side="right", padx=10, pady=10)
    tk.Button(frame, text="Resolver Pendência", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=consultar_usuario_admissao).pack(side="right", padx=10, pady=10)
def relatorio_desligamento():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Relatório Geral de Desligados", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de busca
    filtro_var = tk.StringVar()
    filtro_entry = tk.Entry(frame, textvariable=filtro_var, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    filtro_entry.pack(pady=10, ipady=5)
    
    
    def aplicar_filtro():
        query = filtro_var.get().lower()
        tree.delete(*tree.get_children())
        for _, row in df.iterrows():
            if any(query in str(row[col]).lower() for col in tree['columns']):
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
    
    tk.Button(frame, text="Filtrar", font=("Arial", 12, "bold"), width=20, height=2, bg="#008F68", fg="white", relief="ridge", bd=3, command=aplicar_filtro).pack(pady=10)

    
    # Criando um Frame para conter a Treeview e a Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    # Adicionando Scrollbars horizontal e vertical
    tree_scroll_y = tk.Scrollbar(tree_frame, orient="vertical")
    tree_scroll_y.pack(side="right", fill="y")

    tree_scroll_x = tk.Scrollbar(tree_frame, orient="horizontal")
    tree_scroll_x.pack(side="bottom", fill="x")

    colunas = ["Funcionário","Inativar SISBR", "Bloquear Sipag","Excluir Consórcio","Bloquear Prognum","Redirecionar e-mail","Backup e-mail", "Remover grupos e-mail", "Remover sharepoint", "Bloquear e-mail",
               "Excluir operador topdesk", "Inativar e mover AD","Inativar biometria","Desativar VPN","Comunicar CCS","Observações","Responsável"]
    
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=15, 
                        yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=120, anchor="center")

    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    tree.pack(expand=True, fill="both")
   
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    global df
    if os.path.exists(arquivo):
        df = pd.read_excel(arquivo, sheet_name="Desligados", dtype=str).fillna("")
        
        if not df.empty:
            for _, row in df.iterrows():
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
        else:
            messagebox.showinfo("Relatório", "Nenhum funcionário cadastrado.")
    else:
        messagebox.showerror("Erro", "Arquivo não encontrado.")

    tk.Button(frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=tela_inicial).pack(side="right", padx=10, pady=10)
    tk.Button(frame, text="Resolver Pendência", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=consultar_usuario_desligado).pack(side="right", padx=10, pady=10)

# Caminho do arquivo Excel
CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
ABA_ORIGEM = "Ativos"
ABA_DESTINO = "Inativos"
ABA_CHECKLIST_ADM = "Admissao"
ABA_CHECKLIST = "Desligados"

# Função para salvar checklist
def salvar_checklist(nome):
    responsavel = responsavel_var.get() .strip()    
    observacoes = text_observacoes.get("1.0", "end").strip()
    if responsavel.strip() == "" or responsavel.strip().lower() == "selecione...":
        messagebox.showwarning("Aviso", "Selecione o responsável pelo checklist.")
        return
    # Cria dicionário com o nome do funcionário e os status das tarefas
    linha = {"Funcionário": nome}
    for item in checklist_items:
        linha[item] = "Concluido" if checks[item].get() else "Pendente"
    linha["Observações"] = observacoes
    linha["Responsável"] = responsavel
        
    df_novo = pd.DataFrame([linha])  # DataFrame com uma linha
    if os.path.exists(CAMINHO_ARQUIVO):
        try:
            # Lê o arquivo existente e adiciona a nova linha
            with pd.ExcelWriter(CAMINHO_ARQUIVO, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                workbook = writer.book
                if ABA_CHECKLIST in workbook.sheetnames:
                    sheet = workbook[ABA_CHECKLIST]
                    existing_data = pd.read_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST)
                    df_final = pd.concat([existing_data, df_novo], ignore_index=True)
                    workbook.remove(sheet)
                    df_final.to_excel(writer, sheet_name=ABA_CHECKLIST, index=False)
                else:
                    df_novo.to_excel(writer, sheet_name=ABA_CHECKLIST, index=False)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar checklist: {str(e)}")
    else:
        df_novo.to_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST, index=False)

# Função que move o funcionário da aba Funcionarios para Inativos
def mover_para_inativos(nome_funcionario):
    if not os.path.exists(CAMINHO_ARQUIVO):
        messagebox.showerror("Erro", f"Arquivo '{CAMINHO_ARQUIVO}' não encontrado.")
        return

    wb = load_workbook(CAMINHO_ARQUIVO)
    if ABA_ORIGEM not in wb.sheetnames:
        messagebox.showerror("Erro", f"Aba '{ABA_ORIGEM}' não encontrada.")
        return

    ws_origem = wb[ABA_ORIGEM]
    dados = list(ws_origem.values)
    header = dados[0]
    linhas = dados[1:]

    linha_funcionario = None
    for i, linha in enumerate(linhas, start=2):  # começa da linha 2
        if nome_funcionario.lower() in str(linha[1]).lower():  # Coluna 2 (índice 1) = Nome
            linha_funcionario = (i, linha)
            break

    if not linha_funcionario:
        messagebox.showwarning("Atenção", f"Funcionário '{nome_funcionario}' não encontrado na aba '{ABA_ORIGEM}'.")
        return

    # Criar aba Inativos se não existir
    if ABA_DESTINO not in wb.sheetnames:
        ws_destino = wb.create_sheet(ABA_DESTINO)
        ws_destino.append(header)
    else:
        ws_destino = wb[ABA_DESTINO]

    # Copia e remove
    ws_destino.append(linha_funcionario[1])
    ws_origem.delete_rows(linha_funcionario[0], 1)

    wb.save(CAMINHO_ARQUIVO)
    messagebox.showinfo("Sucesso", f"Funcionário '{nome_funcionario}' movido para a aba'Inativos'.")
    tela_inicial()


# Função para salvar checklist
def salvar_checklist_a(observacoesa):
    #responsavel = responsavel_var.get() .strip()    
    observacoesa = text_observacoesa.get("1.0", "end").strip()
    #if responsavel.strip() == "" or responsavel.strip().lower() == "selecione...":
     #   messagebox.showwarning("Aviso", "Selecione o responsável pelo checklist.")
      #  return
    # Cria dicionário com o nome do funcionário e os status das tarefas
    linha = {"Funcionário": observacoesa}
    for item in checklist_items:
        linha[item] = "Concluido" if checks[item].get() else "Pendente"
    #linha["Nomess"] = observacoesa
    #linha["Responsável"] = responsavel
        
    df_novo = pd.DataFrame([linha])  # DataFrame com uma linha
    if os.path.exists(CAMINHO_ARQUIVO):
        try:
            # Lê o arquivo existente e adiciona a nova linha
            with pd.ExcelWriter(CAMINHO_ARQUIVO, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                workbook = writer.book
                if ABA_CHECKLIST_ADM in workbook.sheetnames:
                    sheet = workbook[ABA_CHECKLIST_ADM]
                    existing_data = pd.read_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST_ADM)
                    df_final = pd.concat([existing_data, df_novo], ignore_index=True)
                    workbook.remove(sheet)
                    df_final.to_excel(writer, sheet_name=ABA_CHECKLIST_ADM, index=False)
                else:
                    df_novo.to_excel(writer, sheet_name=ABA_CHECKLIST_ADM, index=False)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar checklist: {str(e)}")
    else:
        df_novo.to_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST_ADM, index=False)
        
    messagebox.showinfo("Sucesso","Checklist salvo com sucesso!")
    tela_inicial()


# Função principal da tela de checklist
def checklist_desligamento():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Checklist de Desligamento", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")

    # Campo de pesquisa (nome, CPF, e-mail ou login)
    global entry_pesquisa
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa = tk.Entry(linha_pesquisa, font=("Arial", 12), width=42)
    entry_pesquisa.pack(side="left")
    entry_pesquisa.bind("<Return>", lambda event: buscar_funcionario_checklist())
          
    # Campo de exibição do nome encontrado (somente leitura)
    global entry_nome
    entry_nome = tk.Entry(frame, font=("Arial", 12), width=50, state="readonly")
    entry_nome.pack(pady=5)

    global checklist_items, checks
    checklist_items = [
        "Inativar SISBR",
        "Bloquear Sipag",
        "Excluir Consórcio",
        "Bloquear Prognum",
        "Redirecionar e-mail",
        "Backup e-mail",
        "Remover lista distribuição",
        "Remover grupos e-mail",
        "Remover sharepoint",
        "Bloquear e-mail",
        "Excluir operador topdesk",
        "Inativar e mover AD",
        "Inativar biometria",
        "Desativar VPN",
        "Comunicar CCS",        
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Tarefas", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)
    metade = len(checklist_items) // 2 + len(checklist_items) % 2  # Divide a lista

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Frame para responsável + observações lado a lado
    info_frame = tk.Frame(frame, bg="#003D1F")
    info_frame.pack(pady=10)

    # Responsável pelo checklist (label + combobox na mesma linha)
    global responsavel_var
    tk.Label(info_frame, text="Responsável:", font=("Arial", 14), bg="#003D1F", fg="white").grid(row=0, column=0, sticky="e", padx=(0,10))
    responsavel_var = tk.StringVar()
    combo_responsavel = ttk.Combobox(info_frame, textvariable=responsavel_var, values=["Denilson Oliveira", "Paulo Roberto", " Samuel Molina", "Reinaldo Camilo"], font=("Arial", 12), state="readonly", width=30)
    combo_responsavel.grid(row=0, column=1, sticky="w")
    combo_responsavel.set("Selecione...")

    # Observações (label + text box na mesma linha)
    global text_observacoes
    tk.Label(info_frame, text="Observações:", font=("Arial", 14), bg="#003D1F", fg="white").grid(row=1, column=0, sticky="ne", padx=(0,10), pady=(10, 0))
    text_observacoes = tk.Text(info_frame, width=55, height=1, font=("Arial", 12), wrap="word", bd=3, relief="ridge")
    text_observacoes.grid(row=1, column=1, sticky="w", pady=(10, 0))
        
    def ao_clicar_salvar():
        nome = entry_nome.get().strip()
        if not nome:
            messagebox.showwarning("Atenção", "Por favor, digite o nome do funcionário.")
            return
        salvar_checklist(nome)
        mover_para_inativos(nome)
    # Frame para botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)
    
    botoes = [
       # ("Buscar", buscar, "#008F68"),
        ("Salvar", ao_clicar_salvar, "#00B386"),
        ("Relatório/Pendências", relatorio_desligamento, "#00997B"),
        ("Voltar", tela_inicial, "#00997B")
    ]
    
    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)


def salvar_dados(entries):
    dados = {campo: entry.get() for campo, entry in entries.items()}
    
    if any(not valor for valor in dados.values()):
        messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
        return
    
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    salvar_em_excel(dados, arquivo)
    messagebox.showinfo("Sucesso", "Funcionário cadastrado com sucesso!")
    tela_inicial()

def salvar_em_excel(dados, arquivo):
    if os.path.exists(arquivo):
        df_existente = pd.read_excel(arquivo, dtype=str)
        df = pd.concat([df_existente, pd.DataFrame([dados])], ignore_index=True)
    else:
        df = pd.DataFrame([dados])
    
    df.to_excel(arquivo, index=False)
    
    wb = load_workbook(arquivo)
    ws = wb.active
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    fill1 = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill1 if i % 2 == 0 else fill2
        for cell in row:
            cell.fill = fill
    
    wb.save(arquivo)
def buscar_funcionario_checklist():
    termo_busca = entry_pesquisa.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite um nome, CPF, e-mail ou login para buscar.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_funcionarios = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Ativos")

        # Faz busca em várias colunas
        encontrados = df_funcionarios[
            df_funcionarios.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Erro", f"Nenhum funcionário encontrado com o termo '{termo_busca}'.")
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Nome"].tolist())
            messagebox.showerror("Erro", f"Mais de um funcionário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
        else:
            funcionario = encontrados.iloc[0]
            nome_encontrado = funcionario["Nome"]

            # Exibe no campo readonly
            entry_nome.config(state="normal")
            entry_nome.delete(0, tk.END)
            entry_nome.insert(0, nome_encontrado)
            entry_nome.config(state="readonly")

            messagebox.showinfo("Sucesso", f"Funcionário encontrado: {nome_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos funcionários:\n{e}")
# Variáveis globais
df_inativos_global = None
usuario_index_inativo = None

def normalizar_texto(texto):
    """Remove acentos, espaços extras e transforma em minúsculas para facilitar a comparação."""
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto

def consultar_usuario_desligado():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Consulta de Usuários Desligados", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")

    # Campo de pesquisa
    global entry_pesquisa_inativo
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite nome:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa_inativo = tk.Entry(linha_pesquisa, font=("Arial", 12), width=42)
    entry_pesquisa_inativo.pack(side="left")
    entry_pesquisa_inativo.bind("<Return>", lambda event: buscar_usuario_inativo())

    # Campo para mostrar o nome encontrado
    global entry_nome_inativo
    entry_nome_inativo = tk.Entry(frame, font=("Arial", 12), width=50, state="readonly")
    entry_nome_inativo.pack(pady=10)

    # Checklists
    global checklist_items, checks
    checklist_items = [
        "Inativar SISBR",
        "Bloquear Sipag",
        "Excluir Consórcio",
        "Bloquear Prognum",
        "Redirecionar e-mail",
        "Backup e-mail",
        "Remover lista distribuição",
        "Remover grupos e-mail",
        "Remover sharepoint",
        "Bloquear e-mail",
        "Excluir operador topdesk",
        "Inativar e mover AD",
        "Inativar biometria",
        "Desativar VPN",
        "Comunicar CCS",
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Checklist do Usuário", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)

    metade = len(checklist_items) // 2 + len(checklist_items) % 2

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    tk.Button(botoes_frame, text="Buscar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=buscar_usuario_inativo).grid(row=0, column=0, padx=10)
    tk.Button(botoes_frame, text="Salvar Checklist", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=salvar_checklist_inativo).grid(row=0, column=1, padx=10)
    tk.Button(botoes_frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=tela_inicial).grid(row=0, column=2, padx=10)

def buscar_usuario_inativo():
    global df_inativos_global, usuario_index_inativo

    termo_busca = entry_pesquisa_inativo.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite o nome.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_inativos = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Desligados")
        df_inativos_global = df_inativos

        colunas_map = {normalizar_texto(col): col for col in df_inativos.columns}

        encontrados = df_inativos[
            df_inativos.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Não encontrado", f"Nenhum usuário inativo encontrado com o termo '{termo_busca}'.")
            return
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Nome"].tolist())
            messagebox.showerror("Múltiplos resultados", f"Mais de um usuário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
            return
        else:
            usuario = encontrados.iloc[0]
            usuario_index_inativo = encontrados.index[0]

            nome_encontrado = usuario["Funcionário"]
            entry_nome_inativo.config(state="normal")
            entry_nome_inativo.delete(0, tk.END)
            entry_nome_inativo.insert(0, nome_encontrado)
            entry_nome_inativo.config(state="readonly")

            # Atualizar os checkboxes
            for item in checklist_items:
                nome_normalizado = normalizar_texto(item)
                coluna_correspondente = colunas_map.get(nome_normalizado)

                if coluna_correspondente:
                    status = str(usuario.get(coluna_correspondente, "")).strip().lower()
                    checks[item].set(status == "concluido")
                else:
                    checks[item].set(False)

            messagebox.showinfo("Encontrado", f"Usuário desligado encontrado: {nome_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos desligados:\n{e}")

def salvar_checklist_inativo():
    try:
        CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

        if df_inativos_global is None or usuario_index_inativo is None:
            messagebox.showerror("Erro", "Nenhum usuário inativo selecionado para salvar.")
            return

        for item in checklist_items:
            status = "Concluido" if checks[item].get() else "Pendente"
            if item not in df_inativos_global.columns:
                df_inativos_global[item] = ""  # Cria a coluna se não existir
            df_inativos_global.at[usuario_index_inativo, item] = status

        # Salva de volta somente a aba "Inativos"
        with pd.ExcelWriter(CAMINHO_ARQUIVO, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            df_inativos_global.to_excel(writer, sheet_name="Desligados", index=False)

        messagebox.showinfo("Sucesso", "Checklist salvo com sucesso!")
        tela_inicial()

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar o checklist:\n{e}")

def checklist_admissao():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Checklist de Admissão", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")
    info_frame = tk.Frame(frame, bg="#003D1F")
    info_frame.pack(pady=10)
    global text_observacoesa
    tk.Label(info_frame, text="Digite o nome completo:", font=("Arial", 14), bg="#003D1F", fg="white").grid(row=0, column=0, sticky="we", padx=(0,10), pady=(10, 0))
    text_observacoesa = tk.Text(info_frame, width=55, height=1, font=("Arial", 12), wrap="word", bd=3, relief="ridge")
    text_observacoesa.grid(row=1, column=0, sticky="w", pady=(5, 0))
     

    global checklist_items, checks
    checklist_items = [
        "Criar E-mail",
        "Criar AD",
        "Criar SISBR",
        "Criar SIPAG",
        "Criar ASSINATURA",
        "Ramal WEBEX",
        "Portal CCS A/O",
        "CX compartilhada/lista dpt",
        "Colocar lista3246.todos",
        "Acessos Sisbr para central",
        "Grupos outlook",
        "Termo notebook",
        "Cadastrar biometria",        
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Tarefas", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)
    metade = len(checklist_items) // 2 + len(checklist_items) % 2  # Divide a lista

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Frame para responsável + Nomess lado a lado
    info_frame = tk.Frame(frame, bg="#003D1F")
    info_frame.pack(pady=10)
    
       
    def ao_clicar_salvar():
        observacoesa = text_observacoesa.get("1.0", "end").strip()  # <-- Captura o texto digitado
        if not observacoesa:
            messagebox.showwarning("Atenção", "Por favor, digite o nome do funcionário.")
            return
        salvar_checklist_a(observacoesa)
    # Frame para botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)
    
    botoes = [
       # ("Buscar", buscar, "#008F68"),
        ("Salvar", ao_clicar_salvar, "#00B386"),
        ("Relatório/Pendências", relatorio_admissao, "#00997B"),
        ("Voltar", tela_inicial, "#00997B")
    ]
    
    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)


def salvar_dados(entries):
    dados = {campo: entry.get() for campo, entry in entries.items()}
    
    if any(not valor for valor in dados.values()):
        messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
        return
    
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    salvar_em_excel(dados, arquivo)
    messagebox.showinfo("Sucesso", "Funcionário cadastrado com sucesso!")
    tela_inicial()

def salvar_em_excel(dados, arquivo):
    if os.path.exists(arquivo):
        df_existente = pd.read_excel(arquivo, dtype=str)
        df = pd.concat([df_existente, pd.DataFrame([dados])], ignore_index=True)
    else:
        df = pd.DataFrame([dados])
    
    df.to_excel(arquivo, index=False)
    
    wb = load_workbook(arquivo)
    ws = wb.active
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    fill1 = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill1 if i % 2 == 0 else fill2
        for cell in row:
            cell.fill = fill
    
    wb.save(arquivo)
# Variáveis globais
df_admissaos_global = None
usuario_index_admissao = None

def normalizar_texto(texto):
    """Remove acentos, espaços extras e transforma em minúsculas para facilitar a comparação."""
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto

def consultar_usuario_admissao():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Consulta de Usuários Admitidos", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")

    # Campo de pesquisa
    global entry_pesquisa_admissao
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite nome:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa_admissao = tk.Entry(linha_pesquisa, font=("Arial", 12), width=42)
    entry_pesquisa_admissao.pack(side="left")
    entry_pesquisa_admissao.bind("<Return>", lambda event: buscar_usuario_admissao())

    # Campo para mostrar o nome encontrado
    global entry_nome_admissao
    entry_nome_admissao = tk.Entry(frame, font=("Arial", 12), width=50, state="readonly")
    entry_nome_admissao.pack(pady=10)

    # Checklists
    global checklist_items, checks
    checklist_items = [
        "Criar E-mail",
        "Criar AD",
        "Criar SISBR",
        "Criar SIPAG",
        "Criar ASSINATURA",
        "Ramal WEBEX",
        "Portal CCS A/O",
        "CX compartilhada/lista dpt",
        "Colocar lista3246.todos",
        "Acessos Sisbr para central",
        "Grupos outlook",
        "Termo notebook",
        "Cadastrar biometria",
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Checklist do Usuário", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)

    metade = len(checklist_items) // 2 + len(checklist_items) % 2

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    #tk.Button(botoes_frame, text="Buscar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=buscar_usuario_admissao).grid(row=0, column=0, padx=10)
    tk.Button(botoes_frame, text="Salvar Checklist", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=salvar_checklist_a_admissao).grid(row=0, column=1, padx=10)
    tk.Button(botoes_frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=tela_inicial).grid(row=0, column=2, padx=10)

def buscar_usuario_admissao():
    global df_admissaos_global, usuario_index_admissao

    termo_busca = entry_pesquisa_admissao.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite o nome.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_admissaos = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Admissao")
        df_admissaos_global = df_admissaos

        colunas_map = {normalizar_texto(col): col for col in df_admissaos.columns}

        encontrados = df_admissaos[
            df_admissaos.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Não encontrado", f"Nenhum usuário admitido  encontrado com o termo '{termo_busca}'.")
            return
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Nome"].tolist())
            messagebox.showerror("Múltiplos resultados", f"Mais de um usuário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
            return
        else:
            usuario = encontrados.iloc[0]
            usuario_index_admissao = encontrados.index[0]

            nome_encontrado = usuario["Funcionário"]
            entry_nome_admissao.config(state="normal")
            entry_nome_admissao.delete(0, tk.END)
            entry_nome_admissao.insert(0, nome_encontrado)
            entry_nome_admissao.config(state="readonly")

            # Atualizar os checkboxes
            for item in checklist_items:
                nome_normalizado = normalizar_texto(item)
                coluna_correspondente = colunas_map.get(nome_normalizado)

                if coluna_correspondente:
                    status = str(usuario.get(coluna_correspondente, "")).strip().lower()
                    checks[item].set(status == "concluido")
                else:
                    checks[item].set(False)

            messagebox.showinfo("Encontrado", f"Usuário admitido encontrado: {nome_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos admitidos:\n{e}")
    
def salvar_checklist_a_admissao():
    try:
        CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

        if df_admissaos_global is None or usuario_index_admissao is None:
            messagebox.showerror("Erro", "Nenhum usuário admitido selecionado para salvar.")
            return

        for item in checklist_items:
            status = "Concluido" if checks[item].get() else "Pendente"
            if item not in df_admissaos_global.columns:
                df_admissaos_global[item] = ""  # Cria a coluna se não existir
            df_admissaos_global.at[usuario_index_admissao, item] = status

        # Salva de volta somente a aba "Inativos"
        with pd.ExcelWriter(CAMINHO_ARQUIVO, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            df_admissaos_global.to_excel(writer, sheet_name="Admissao", index=False)

        messagebox.showinfo("Sucesso", "Checklist salvo com sucesso!")
        tela_inicial()

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar o checklist:\n{e}")
    
def buscar_funcionario_grupo():
    termo_busca = entry_pesquisa.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite um nome, CPF, e-mail ou login para buscar.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_funcionarios = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Ativos")

        # Faz busca em várias colunas
        encontrados = df_funcionarios[
            df_funcionarios.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Erro", f"Nenhum funcionário encontrado com o termo '{termo_busca}'.")
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Login"].tolist())
            messagebox.showerror("Erro", f"Mais de um funcionário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
        else:
            funcionario = encontrados.iloc[0]
            login_encontrado = funcionario["Login"]

            # Exibe no campo readonly
            entry_login.config(state="normal")
            entry_login.delete(0, tk.END)
            entry_login.insert(0, login_encontrado)
            entry_login.config(state="readonly")

            messagebox.showinfo("Sucesso", f"Funcionário encontrado: {login_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos funcionários:\n{e}")


# Configuração inicial da janela
root = tk.Tk()
root.title("Gestão de Funcionários")
root.geometry("1100x800")

tela_inicial()

root.mainloop()
<<<<<<< HEAD
=======
=======
import pandas as pd
import os
import tkinter as tk
from tkinter import PhotoImage, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from concurrent.futures import ThreadPoolExecutor
import webbrowser
import unicodedata
import re
import subprocess
import platform
import threading
import socket
import time
import csv



def limpar_tela():
    for widget in root.winfo_children():
        widget.destroy()

import tkinter as tk

def tela_inicial():
    limpar_tela()
    
    root.configure(bg="#003D1F")  # Fundo verde ainda mais escuro para maior contraste

       
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)
    
    # Cabeçalho
    header = tk.Frame(frame, bg="#026440", pady=10)  # Verdee intermediário para destaque
    header.pack(fill="x")
    
    label_titulo = tk.Label(header, text="Gestão de acesso Tecnologia - 3246", font=("Arial", 22, "bold"), bg="#026440", fg="white")
    label_titulo.pack(pady=10)
    
    # Área de Botões
    main_frame = tk.Frame(frame, bg="#003D1F")
    main_frame.pack(pady=20)
    
    botoes = [
        ("Colaboradores", cadastrar_ou_alterar_funcionario, "#008F68"),
        ("Acessar UEM", pesquisar_funcionario, "#00B386"),
        ("Checklist Admissão", checklist_admissao, "#00997B"),
        ("Checklist Demissão", checklist_desligamento, "#00997B"),
        ("Grupo de Acesso", grupo_acesso, "#008386"),
        ("Ip Scanner", varredura_ip, "#008386"),
        ("Relatórios", relatorio_geral, "#66CDAA")
    ]

    # Criar botões em 2 colunas
    for i, (text, command, color) in enumerate(botoes):
        row = i // 2   # Linha (vai mudando a cada 2 botões)
        col = i % 2    # Coluna (0 ou 1)
        
        tk.Button(
            main_frame,
            text=text,
            font=("Arial", 14, "bold"),
            width=25,
            height=2,
            bg=color,
            fg="white",
            relief="ridge",
            bd=3,
            command=command
        ).grid(row=row, column=col, padx=15, pady=15)

    # Centralizar as colunas
    main_frame.grid_columnconfigure(0, weight=1)
    main_frame.grid_columnconfigure(1, weight=1)
    
    # Rodapé
    footer = tk.Frame(frame, bg="#026440", pady=5)
    footer.pack(fill="x", side="bottom")
    
    label_footer = tk.Label(footer, text="Sicoob Credseguro - 2025", font=("Arial", 10), bg="#026440", fg="white")
    label_footer.pack()

def cadastrar_ou_alterar_funcionario():
    limpar_tela()

    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Cadastro/Alteração de Funcionário", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    tk.Label(frame, text="Digite nome, CPF, login ou e-mail para buscar:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(pady=10)
    entry_pesquisa = tk.Entry(frame, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    entry_pesquisa.pack(pady=10, ipady=5)
    entry_pesquisa.bind("<Return>", lambda event: buscar())

    form_frame = tk.Frame(frame, bg="#003D1F")
    form_frame.pack(pady=10)

    campos = ["CPF", "Nome", "Computador", "AD", "Login", "E-mail", "Consórcio", "Telefone", "IP","MAC" , "UEM"]
    entries = {}

    # Funções adicionais
    IP_FIXO = "10.201."

    DOMINIO_FIXO = ""

    def verificar_duplicidade(event=None):
        caminho_arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")  # ajuste se necessário

        if not os.path.exists(caminho_arquivo):
            return

        try:
            df = pd.read_excel(caminho_arquivo)

            # Normalizar colunas (garantir que existam)
            colunas_esperadas = ["Nome", "CPF", "IP"]
            for col in colunas_esperadas:
                if col not in df.columns:
                    return

        except Exception as e:
            messagebox.showerror("Erro ao verificar duplicidade", f"Erro ao ler o arquivo: {e}")
    def remover_acentos(texto):
        return ''.join(
            c for c in unicodedata.normalize('NFKD', texto)
            if not unicodedata.combining(c)
        )
    def preencher_ad(event=None):
        nome_completo = entries["Nome"].get().strip()

        if not nome_completo:
            return

        nome_sem_acentos = remover_acentos(nome_completo)
        nomes = re.findall(r'\b\w+\b', nome_sem_acentos.lower())  # pega só palavras com letras/números

        if len(nomes) >= 2:
            primeiro = nomes[0]
            ultimo = nomes[-1]
            ad_formatado = f"{primeiro}.{ultimo}"
        elif nomes:
            ad_formatado = nomes[0]
        else:
            ad_formatado = ""

        entries["AD"].delete(0, tk.END)
        entries["AD"].insert(0, ad_formatado)
    def aplicar_mascara_mac(event):
        entry = entries["MAC"]
        texto = entry.get().upper()
        texto_hex = ''.join(filter(lambda c: c in "0123456789ABCDEF", texto))[:12]

        formatado = ":".join(texto_hex[i:i+2] for i in range(0, len(texto_hex), 2))

        entry.delete(0, tk.END)
        entry.insert(0, formatado)
    def formatar_nome(event):
        entry = entries["Nome"]
        texto = entry.get()
        cursor_pos = entry.index(tk.INSERT)

        # Não faz nada se campo estiver vazio
        if not texto:
            return

        # Preserva espaços múltiplos entre palavras enquanto digita
        palavras = texto.split(' ')
        palavras_formatadas = [p.capitalize() if p else '' for p in palavras]
        texto_corrigido = ' '.join(palavras_formatadas)

        # Atualiza o campo só se houver mudança
        if texto != texto_corrigido:
           entry.delete(0, tk.END)
           entry.insert(0, texto_corrigido)
           entry.icursor(min(cursor_pos, len(texto_corrigido)))

    def limpar_nome(event):
        entry = entries["Nome"]
        texto = entry.get()
        texto_filtrado = ''.join(
            char for char in texto if char.isalpha() or char.isspace() or unicodedata.category(char).startswith('L')
        )
        entry.delete(0, tk.END)
        entry.insert(0, texto_filtrado)
    
    def formatar_maiusculo(event):
        entry = event.widget
        texto = entry.get()
        entry.delete(0, tk.END)
        entry.insert(0, texto.upper())
    def aplicar_mascara_telefone(event=None):
        entry = event.widget
        texto = entry.get()
        numeros = ''.join(filter(str.isdigit, texto))[:11]

        if len(numeros) <= 2:
            formatado = f"({numeros}"
        elif len(numeros) <= 7:
            formatado = f"({numeros[:2]}){numeros[2:]}"
        else:
            formatado = f"({numeros[:2]}){numeros[2:7]}-{numeros[7:]}"

        entry.delete(0, tk.END)
        entry.insert(0, formatado)
    
    def aplicar_mascara_cpf(event):
        texto = entries["CPF"].get()
        texto_numerico = ''.join(filter(str.isdigit, texto))[:11]
        novo_texto = ""

        for i in range(len(texto_numerico)):
            if i == 3 or i == 6:
                novo_texto += "."
            elif i == 9:
                novo_texto += "-"
            novo_texto += texto_numerico[i]

        entries["CPF"].delete(0, tk.END)
        entries["CPF"].insert(0, novo_texto)

    def cpf_valido(cpf):
        cpf = ''.join(filter(str.isdigit, cpf))
        if len(cpf) != 11 or cpf == cpf[0] * 11:
            return False
        soma1 = sum(int(cpf[i]) * (10 - i) for i in range(9))
        digito1 = (soma1 * 10 % 11) % 10
        soma2 = sum(int(cpf[i]) * (11 - i) for i in range(10))
        digito2 = (soma2 * 10 % 11) % 10
        return cpf[-2:] == f"{digito1}{digito2}"

    def validar_cpf(event):
        cpf_texto = entries["CPF"].get()
        if not cpf_valido(cpf_texto):
            messagebox.showerror("CPF Inválido", "Digite um CPF válido.")
            entries["CPF"].focus_set()

    
    def manter_sufixo_email(event):
        texto = entries["E-mail"].get()
        if DOMINIO_FIXO in texto:
            prefixo = texto.split(DOMINIO_FIXO)[0]
        else:
            prefixo = texto.replace(DOMINIO_FIXO, ". ")
        novo_email = prefixo + DOMINIO_FIXO
        entries["E-mail"].delete(0, tk.END)
        entries["E-mail"].insert(0, novo_email)
        entries["E-mail"].icursor(len(prefixo))

    def validar_email(event):
        email = entries["E-mail"].get()
        if not email or email == DOMINIO_FIXO:
            messagebox.showerror("E-mail inválido", "Não precisa digitar '@sicoob.com.br'")
            entries["E-mail"].focus_set()

    def manter_prefixo_ip(event):
        texto = entries["IP"].get()
        sufixo = texto.replace(IP_FIXO, "")
        numeros = ''.join(filter(str.isdigit, sufixo))[:6]

        bloco1 = numeros[:2]
        bloco2 = numeros[2:]

        novo_ip = IP_FIXO
        if bloco1:
            novo_ip += bloco1
        if bloco2:
            novo_ip += "." + bloco2

        entries["IP"].delete(0, tk.END)
        entries["IP"].insert(0, novo_ip)
        entries["IP"].icursor(tk.END)

    def validar_ip(event):
        ip = entries["IP"].get()
        partes = ip.split(".")
        if len(partes) != 4:
            messagebox.showerror("IP Inválido", "O IP deve ter o formato 10.201.xx.xxx")
            entries["IP"].focus_set()
            return
        try:
            for parte in partes:
                if not parte.isdigit() or not (0 <= int(parte) <= 255):
                    raise ValueError
        except:
            messagebox.showerror("IP Inválido", "Cada parte do IP deve estar entre 0 e 255.")
            entries["IP"].focus_set()
    
    def somente_letras(event):
        texto = entries["Nome"].get()

        # Permitir letras, acentos e espaços
        texto_filtrado = ''.join(char for char in texto if (
        char.isalpha() or char.isspace() or unicodedata.category(char).startswith('L')
    ))

        # Atualiza o campo com apenas caracteres permitidos
        if texto != texto_filtrado:
            entries["Nome"].delete(0, tk.END)
            entries["Nome"].insert(0, texto_filtrado)
    
    for i, campo in enumerate(campos):
        row, col = divmod(i, 2)
        tk.Label(
            form_frame, text=f"{campo}:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w"
        ).grid(row=row, column=col * 2, sticky="w", padx=10, pady=5)

        entry = tk.Entry(
            form_frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F"
        )
        entry.grid(row=row, column=col * 2 + 1, padx=10, pady=1, ipady=5, sticky="w")
        entries[campo] = entry

        if campo == "CPF":
            entry.bind("<KeyRelease>", aplicar_mascara_cpf)
            entry.bind("<FocusOut>", validar_cpf)
            entry.bind("<FocusOut>", verificar_duplicidade)

        if campo == "E-mail":
            #entry.insert(0, DOMINIO_FIXO)
            #entry.bind("<KeyRelease>", manter_sufixo_email)
            #entry.bind("<FocusOut>", validar_email)
            pass

        if campo == "IP":
            entry.insert(0, IP_FIXO)
            entry.bind("<KeyRelease>", manter_prefixo_ip)
            entry.bind("<FocusOut>", validar_ip)
            entry.bind("<FocusOut>", verificar_duplicidade)

        if campo == "Nome" :
            entry.bind("<KeyRelease>", formatar_nome)
            entry.bind("<FocusOut>", limpar_nome)
            entry.bind("<KeyRelease>", preencher_ad)
            entry.bind("<FocusOut>", verificar_duplicidade)
                    
        if campo == "Consórcio":
            entry.insert(0, "Não tem")

        if campo == "Login" :
            entry.insert(0, "")    

        if campo in ["Login", "Computador"]:
            entry.bind("<KeyRelease>", formatar_maiusculo)
            entry.bind("<FocusOut>", formatar_maiusculo)
            entry.bind("<FocusOut>", verificar_duplicidade)
        
        if campo == "Telefone":
            entry.insert(0, "(62)3275-0200")
            entry.bind("<KeyRelease>", aplicar_mascara_telefone)

        if campo == "MAC":
            entry.bind("<KeyRelease>", aplicar_mascara_mac)

                    
    # Base da linha onde os campos adicionais começam
    linha_base = (len(campos) + 1) // 2

    # PA
    pa_var = tk.StringVar()
    valores_pa = [" ", "UAD", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"]
    tk.Label(form_frame, text="PA:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base, column=0, sticky="w", padx=10, pady=5)

    pa_combobox = ttk.Combobox(form_frame, textvariable=pa_var, values=valores_pa, state="readonly", font=("Arial", 12), width=38)
    pa_combobox.grid(row=linha_base, column=1, padx=10, pady=5, ipady=2)
    entries["PA"] = pa_combobox

    # Prognum
    prognum_var = tk.StringVar()
    valores_prognum = [" ", "Sim", "Não"]
    tk.Label(form_frame, text="Prognum:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base, column=2, sticky="w", padx=10, pady=5)

    prognum_combobox = ttk.Combobox(form_frame, textvariable=prognum_var, values=valores_prognum, state="readonly", font=("Arial", 12), width=38)
    prognum_combobox.grid(row=linha_base, column=3, padx=10, pady=5, ipady=2)
    entries["Prognum"] = prognum_combobox

    # Cargo
    cargo_var = tk.StringVar()
    valores_cargo = [" ", "Auxiliar", "Assistente", "Agente de atendimento", "Agente de relacionamento", "Analista", "Caixa", "Conselheiro(a)", "Diretor", "Estagiário", "Especialista", "Gerente de relacionamento", "Gestor", "Jovem aprendiz", "Presidente", "Secretaria", "Superitendente", "Supervisor(a)", "Tesoureiro(a)"]
    tk.Label(form_frame, text="Cargo:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base + 1, column=0, sticky="w", padx=10, pady=5)
 
    cargo_combobox = ttk.Combobox(form_frame, textvariable=cargo_var, values=valores_cargo, state="readonly", font=("Arial", 12), width=38)
    cargo_combobox.grid(row=linha_base + 1, column=1, padx=10, pady=5, ipady=2)
    entries["Cargo"] = cargo_combobox

    # Setor
    setor_var = tk.StringVar()
    valores_setor = [" ", "Administrativo", "Advocacia", "Cadastro", "Crédito", "Cobrança", "Comercial", "Conselho","Controladoria", "Diretoria", "Financeiro", "Financeiro UAD", "Secretária", "Gente e Gestão", "Marketing", "Produtos", "Tecnologia","Superitendencia", "Sustentabilidade", "Riscos e Controles"]
    tk.Label(form_frame, text="Setor:", font=("Arial", 12, "bold"), bg="#003D1F", fg="white", anchor="w")\
    .grid(row=linha_base + 1, column=2, sticky="w", padx=10, pady=5)

    setor_combobox = ttk.Combobox(form_frame, textvariable=setor_var, values=valores_setor, state="readonly", font=("Arial", 12), width=38)
    setor_combobox.grid(row=linha_base + 1, column=3, padx=10, pady=5, ipady=2)
    entries["Setor"] = setor_combobox

    # O metodo buscar reflete na tela de cadastro de novo colaborador
    def buscar():        
        termo = entry_pesquisa.get().strip().lower()
        if not termo:
            messagebox.showerror("Erro", "Digite um termo para pesquisa!")
            
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):

            df = pd.read_excel(arquivo, dtype=str).fillna("")
            resultado = df[df.apply(lambda row: any(termo in str(valor).lower() for valor in row), axis=1)]

            if not resultado.empty:
                dados = resultado.iloc[0]
               # Preenchendo os campos de texto normais
                for campo in campos:
                    entries[campo].delete(0, tk.END)
                    entries[campo].insert(0, dados[campo])
            # Preenchendo o campo "PA"
                if "PA" in dados and dados["PA"] in valores_pa:
                    pa_var.set(dados["PA"])
                else:
                    pa_var.set(" ")
            # Preenchendo o campo "Cargo"
                if "Cargo" in dados and dados["Cargo"] in valores_cargo:
                    cargo_var.set(dados["Cargo"])
                else:
                    cargo_var.set(" ")

            # Preenchendo o campo "Setor"
                if "Setor" in dados and dados["Setor"] in valores_setor:
                    setor_var.set(dados["Setor"])
                else:
                    setor_var.set(" ")
            # Preenchendo corretamente o campo "Prognum" na Combobox
                if "Prognum" in dados and dados["Prognum"] in ["Sim", "Não"]:
                    prognum_var.set(dados["Prognum"])  # Atualiza a Combobox com o valor correto
                else:
                    prognum_var.set("erro")  # Define "Sim" como padrão caso o valor não seja válido

            else:
                messagebox.showinfo("Pesquisa", "Nenhum funcionário encontrado.")
        else:
             messagebox.showerror("Erro", "Arquivo não encontrado.")

    def excluir_funcionario():
        cpf = entries["CPF"].get().strip()
        cpf_numerico = ''.join(filter(str.isdigit, cpf))

        if not cpf_numerico:
            messagebox.showerror("Erro", "Nenhum CPF informado. Faça a busca de um funcionário primeiro.")
            return

        confirmacao = messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir o funcionário com CPF {cpf}?")
        if not confirmacao:
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):
            df = pd.read_excel(arquivo, dtype=str).fillna("")
            df_filtrado = df[df["CPF"].str.replace(r'\D', '', regex=True) != cpf_numerico]

            if len(df) == len(df_filtrado):
                messagebox.showinfo("Não encontrado", "Funcionário não encontrado para exclusão.")
            else:
                df_filtrado.to_excel(arquivo, index=False)
                messagebox.showinfo("Excluído", "Funcionário excluído com sucesso.")
                tela_inicial()
        else:
            messagebox.showerror("Erro", "Arquivo de funcionários não encontrado.")
    def salvar_alteracoes():
        novo_dados = {campo: entries[campo].get() for campo in campos}
        campos_reordenados = ["CPF", "Nome", "Computador", "AD", "Login", "E-mail", "Prognum", "Consórcio", "PA", "Setor", "Telefone", "IP", "MAC", "Cargo", "UEM"]
        novo_dados["PA"] = pa_var.get()
        novo_dados["Prognum"] = prognum_var.get()
        novo_dados["Cargo"] = cargo_var.get()
        novo_dados["Setor"] = setor_var.get()         

        if any(not valor for valor in novo_dados.values()):
            messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):
            # Lê todas as abas
            all_sheets = pd.read_excel(arquivo, sheet_name=None, dtype=str)
            all_sheets = {nome: df.fillna("") for nome, df in all_sheets.items()}

            nome_aba_principal = "Ativos"  # substitua pelo nome real da aba principal, se necessário
            df = all_sheets.get(nome_aba_principal)

            if df is None:
                messagebox.showerror("Erro", f"A aba '{nome_aba_principal}' não foi encontrada no arquivo!")
                return

            if list(df.columns) != campos_reordenados:
                messagebox.showerror("Erro", "A ordem das colunas na planilha não está correta!")
                return 
            
            index = df[df["CPF"] == novo_dados["CPF"]].index
            if not index.empty:
                df.loc[index[0], :] = [novo_dados[campo] for campo in campos_reordenados]
                all_sheets[nome_aba_principal] = df
            else:
                 # Adiciona novo funcionário
                novo_df = pd.DataFrame([novo_dados], columns=campos_reordenados)
                all_sheets[nome_aba_principal] = pd.concat([df, novo_df], ignore_index=True)

            # Salva todas as abas de volta no Excel
            with pd.ExcelWriter(arquivo, engine='openpyxl', mode='w') as writer:
                for aba, data in all_sheets.items():
                    data.to_excel(writer, sheet_name=aba, index=False)

            messagebox.showinfo("Sucesso", "Alterações salvas com sucesso!")
            tela_inicial()
        else:
            messagebox.showerror("Erro", "Arquivo não encontrado!")

    # Frame para botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)
    
    botoes = [
       # ("Buscar", buscar, "#008F68"),
        ("Salvar", salvar_alteracoes, "#00B386"),
        ("Excluir", excluir_funcionario, "#00B386"),
        ("Voltar", tela_inicial, "#00997B")
    ]
    
    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)


def grupo_acesso():
    limpar_tela()

    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Grupo de acesso", font=("Arial", 18, "bold"),bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de pesquisa (nome, CPF, e-mail ou login)
    global entry_pesquisa
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite o nome", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa = tk.Entry(linha_pesquisa, font=("Arial", 12), width=26)
    entry_pesquisa.pack(side="left")
    entry_pesquisa.bind("<Return>", lambda event: buscar_funcionario_grupo())
   
    global entry_login
    entry_login = tk.Entry(frame, font=("Arial", 12), width=42, state="readonly")
    entry_login.pack(pady=5)

    # Carrega grupos do Excel
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    grupos = {}

    if os.path.exists(arquivo):
        try:
            df = pd.read_excel(arquivo, sheet_name="Grupo")
            for _, row in df.iterrows():
                cargo = str(row[0]).strip()
                grupo = str(row[1]).strip()
                if cargo and grupo:
                    grupos[cargo] = grupo
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler a planilha: {e}")
            return
    else:
        messagebox.showerror("Erro", "Arquivo 'funcionarios.xlsx' não encontrado.")
        return

    if not grupos:
        messagebox.showinfo("Aviso", "Nenhum grupo encontrada na aba 'Grupo'.")
        return

    # Dropdown de cargos
    combo_frame = tk.Frame(frame, bg="#003D1F")
    combo_frame.pack(pady=10)

    tk.Label(combo_frame, text="Escolha a cargo:", font=("Arial", 14, "bold"),  bg="#003D1F", fg="white").pack(side="left", padx=(0,10))

    combo = ttk.Combobox(combo_frame, values=list(grupos.keys()), state="readonly", font=("Arial", 12), width=20)
    combo.pack(side="left")

    # Caixa de texto com scrollbar (tamanho reduzido)
    text_frame = tk.Frame(frame, bg="#003D1F")
    text_frame.pack(pady=10)

    caixa_texto = tk.Text(text_frame, wrap="word", yscrollcommand=lambda *args: scrollbar.set(*args),
                      font=("Arial", 12), bg="#E8F5E9", fg="#003D1F", relief="ridge", bd=3,
                      width=90, height=22)
    caixa_texto.pack(side="left")

    scrollbar = tk.Scrollbar(text_frame, command=caixa_texto.yview)
    scrollbar.pack(side="right", fill="y")

    def mostrar_grupo(event=None):
        cargo = combo.get()
        grupo = grupos.get(cargo, "grupo não encontrado.")
        caixa_texto.delete("1.0", tk.END)
        caixa_texto.insert(tk.END, grupo)

    combo.bind("<<ComboboxSelected>>", mostrar_grupo)

    # Botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    def copiar_texto():
        texto = caixa_texto.get("1.0", tk.END).strip()
        if texto:
            root.clipboard_clear()
            root.clipboard_append(texto)
            messagebox.showinfo("Copiado", "Texto copiado para a área de transferência.")

    def salvar_texto():
        cargo = combo.get()
        novo_grupo = caixa_texto.get("1.0", tk.END).strip()
        if not cargo:
            messagebox.showwarning("Aviso", "Selecione uma cargo antes de salvar.")
            return
        try:
            df = pd.read_excel(arquivo, sheet_name="Grupo")
            cargo_encontrada = False
            for i in range(len(df)):
                if str(df.iloc[i, 0]).strip() == cargo:
                    df.iat[i, 1] = novo_grupo
                    cargo_encontrada = True
                    break
            if cargo_encontrada:
                with pd.ExcelWriter(arquivo, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df.to_excel(writer, sheet_name="Grupo", index=False)
                messagebox.showinfo("Sucesso", "Grupo atualizado com sucesso.")
            else:
                messagebox.showwarning("Aviso", "Cargo não encontrado na planilha.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar: {e}")

    botoes = [
        ("Copiar", copiar_texto, "#026440"),
        ("Salvar", salvar_texto, "#008F68"),
        ("Voltar", tela_inicial, "#00997B")
    ]

    for i, (texto, comando, cor) in enumerate(botoes):
        tk.Button(botoes_frame, text=texto, font=("Arial", 12, "bold"), width=20, height=2,
                  bg=cor, fg="white", relief="ridge", bd=3, command=comando).grid(row=0, column=i, padx=10)
def pesquisar_funcionario():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Acesso Remoto UEM", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    tk.Label(frame, text="Digite nome do usuário:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(pady=10)
    entry_pesquisa = tk.Entry(frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    entry_pesquisa.pack(padx=20, pady=10, ipady=5)

    # Associa a tecla "Enter" à função de busca
    entry_pesquisa.bind("<Return>", lambda event: buscar_uem())

    # Adicionando Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side="right", fill="y")

    colunas = ["Computador", "IP", "Nome", "Setor", "PA", "MAC", "Cargo", "UEM"]
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=10, yscrollcommand=tree_scroll.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=200, anchor="center")

    tree_scroll.config(command=tree.yview)
    tree.pack(expand=True, fill="both")

    def abrir_link(event):
        item = tree.selection()
        if item:
            valores = tree.item(item, "values")
            link = valores[-1]  # Última coluna (UEM)
            if link.startswith("http"):
                webbrowser.open(link)

    tree.bind("<Double-1>", abrir_link)

    def buscar_uem():
        for item in tree.get_children():
            tree.delete(item)

        uem = entry_pesquisa.get().strip().lower()
        if not uem:
            messagebox.showerror("Erro", "Digite um termo para pesquisa!")
            return

        arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
        if os.path.exists(arquivo):
            try:
                # Ler as abas "Ativos" e "PC"
                df_ativos = pd.read_excel(arquivo, sheet_name="Ativos", dtype=str).fillna("")
                df_pc = pd.read_excel(arquivo, sheet_name="PC", dtype=str).fillna("")

                # Filtrar resultados nas duas abas
                resultados_ativos = df_ativos[df_ativos.apply(lambda row: any(uem in str(valor).lower() for valor in row), axis=1)]
                resultados_pc = df_pc[df_pc.apply(lambda row: any(uem in str(valor).lower() for valor in row), axis=1)]

                # Concatenar os resultados
                resultados = pd.concat([resultados_ativos, resultados_pc], ignore_index=True)
                
                if not resultados.empty:
                    for _, row in resultados.iterrows():
                        tree.insert("", "end", values=[row[col] for col in tree['columns']])
                else:
                    messagebox.showinfo("Pesquisa", "Nenhum funcionário encontrado.")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        else:
            messagebox.showerror("Erro", "Arquivo não encontrado.")

    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    botoes = [
        ("Buscar", buscar_uem, "#008F68"),
        ("Voltar", tela_inicial, "#00997B")
    ]

    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)

def ping_ip(ip):
    param = "-n" if platform.system().lower() == "windows" else "-c"
    comando = ["ping", param, "1", ip]
    try:
        subprocess.check_output(comando, stderr=subprocess.DEVNULL, universal_newlines=True, timeout=1)
        return True
    except subprocess.CalledProcessError:
        return False
    except subprocess.TimeoutExpired:
        return False

def obter_hostname(ip):
    try:
        return socket.gethostbyaddr(ip)[0]
    except socket.herror:
        return ""
    
def varredura_ip():
    limpar_tela()

    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Varredura de IP na Rede", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo faixa IP
    instrucao = tk.Label(frame, text="Informe a faixa de IP (ex: 10.201.60.1-10.201.60.254):", font=("Arial", 12), bg="#003D1F", fg="white")
    instrucao.pack(pady=5)
    entry_faixa = tk.Entry(frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9")
    entry_faixa.pack(pady=5, ipady=3)

    # Campo filtro hostname
    filtro_label = tk.Label(frame, text="Filtrar por Hostname (opcional):", font=("Arial", 12), bg="#003D1F", fg="white")
    filtro_label.pack(pady=5)
    entry_filtro = tk.Entry(frame, width=40, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9")
    entry_filtro.pack(pady=5, ipady=3)

    # Barra de progresso
    progress_var = tk.DoubleVar()
    progress = ttk.Progressbar(frame, variable=progress_var, maximum=100)
    progress.pack(fill="x", pady=10)

    # Tabela Treeview
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(fill="both", expand=True, pady=10)

    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side="right", fill="y")

    tree = ttk.Treeview(tree_frame, columns=("IP", "Host"), show="headings", yscrollcommand=tree_scroll.set)
    tree.heading("IP", text="Endereço IP")
    tree.heading("Host", text="Hostname")
    tree.column("IP", width=150)
    tree.column("Host", width=300)
    tree.pack(fill="both", expand=True)

    # Logs detalhados
    log_frame = tk.Frame(frame)
    log_frame.pack(fill="both", expand=False, pady=5)
    
    text_log_scroll = tk.Scrollbar(log_frame)
    text_log_scroll.pack(side="right", fill="y")

    text_log = tk.Text(log_frame, height=10, bg="black", fg="lime", font=("Courier", 10), yscrollcommand=text_log_scroll.set)
    text_log.pack(fill="both", expand=True)

    text_log_scroll.config(command=text_log.yview)

    btn_frame = tk.Frame(frame, bg="#003D1F")
    btn_frame.pack(pady=10)

    resultado = []

    def log(msg):
        text_log.insert("end", msg + "\n")
        text_log.see("end")  # Scroll automático

    def iniciar_varredura():
        faixa = entry_faixa.get().strip()
        filtro = entry_filtro.get().strip().lower()
        for i in tree.get_children():
            tree.delete(i)
        resultado.clear()
        progress_var.set(0)
        text_log.delete("1.0", "end")

        if "-" not in faixa:
            messagebox.showerror("Erro", "Informe a faixa corretamente (ex: 192.168.1.1-192.168.1.254).")
            return

        ip_inicio, ip_fim = faixa.split("-")
        partes_inicio = ip_inicio.strip().split(".")
        partes_fim = ip_fim.strip().split(".")
        if partes_inicio[:-1] != partes_fim[:-1]:
            messagebox.showerror("Erro", "O intervalo deve estar na mesma sub-rede.")
            return

        prefixo = ".".join(partes_inicio[:-1]) + "."
        ini = int(partes_inicio[-1])
        fim = int(partes_fim[-1])
        total = fim - ini + 1

        encontrados = []

        def escanear_ip(i):
            ip = prefixo + str(i)
            log(f"Pinging {ip}...")
            if ping_ip(ip):
                host = obter_hostname(ip)
                log(f"✔ {ip} respondeu - Hostname: {host}")
                if filtro == "" or filtro in host.lower():
                    encontrados.append((ip, host))
                    tree.insert("", "end", values=(ip, host))
            else:
                log(f"✖ {ip} não respondeu")
            progress_var.set((i - ini + 1) / total * 100)

        def escanear():
            btn_iniciar.config(state="disabled")
            btn_exportar.config(state="disabled")
            with ThreadPoolExecutor(max_workers=50) as executor:
                futures = [executor.submit(escanear_ip, i) for i in range(ini, fim + 1)]
                for f in futures:
                    f.result()
            resultado.extend(encontrados)
            log("✅ Varredura finalizada!")
            progress_var.set(100)
            btn_iniciar.config(state="normal")
            btn_exportar.config(state="normal")

        threading.Thread(target=escanear).start()

    def exportar_csv():
        if not resultado:
            messagebox.showinfo("Informação", "Nenhum dado para exportar.")
            return
        nome_arquivo = f"varredura_{time.strftime('%Y%m%d_%H%M%S')}.csv"
        with open(nome_arquivo, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Endereço IP", "Hostname"])
            writer.writerows(resultado)
        messagebox.showinfo("Exportado", f"Arquivo salvo: {nome_arquivo}")

    btn_iniciar = tk.Button(btn_frame, text="Iniciar Varredura", font=("Arial", 12, "bold"), bg="#00B386", fg="white", command=iniciar_varredura)
    btn_iniciar.pack(side="left", padx=5)

    btn_exportar = tk.Button(btn_frame, text="Exportar CSV", font=("Arial", 12, "bold"), bg="#00997B", fg="white", command=exportar_csv, state="disabled")
    btn_exportar.pack(side="left", padx=5)

    btn_voltar = tk.Button(btn_frame, text="Voltar", font=("Arial", 12, "bold"), bg="#008F68", fg="white", command=tela_inicial)
    btn_voltar.pack(side="left", padx=5)

def relatorio_geral():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Relatório Geral de Funcionários", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de busca
    filtro_var = tk.StringVar()
    filtro_entry = tk.Entry(frame, textvariable=filtro_var, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    filtro_entry.pack(pady=10, ipady=5)
    
    
    def aplicar_filtro():
        query = filtro_var.get().lower()
        tree.delete(*tree.get_children())
        for _, row in df.iterrows():
            if any(query in str(row[col]).lower() for col in tree['columns']):
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
    
    tk.Button(frame, text="Filtrar", font=("Arial", 12, "bold"), width=20, height=2, bg="#008F68", fg="white", relief="ridge", bd=3, command=aplicar_filtro).pack(pady=10)

    
    # Criando um Frame para conter a Treeview e a Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    # Adicionando Scrollbars horizontal e vertical
    tree_scroll_y = tk.Scrollbar(tree_frame, orient="vertical")
    tree_scroll_y.pack(side="right", fill="y")

    tree_scroll_x = tk.Scrollbar(tree_frame, orient="horizontal")
    tree_scroll_x.pack(side="bottom", fill="x")

    colunas = ["CPF", "Nome", "Computador", "AD", "Login", "E-mail", "Prognum", "Consórcio", "PA", "Setor", "Cargo", "Telefone", "IP", "MAC"]
    
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=15, 
                        yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=120, anchor="center")

    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    tree.pack(expand=True, fill="both")
   
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    global df
    if os.path.exists(arquivo):
        df = pd.read_excel(arquivo, dtype=str).fillna("")
        
        if not df.empty:
            for _, row in df.iterrows():
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
        else:
            messagebox.showinfo("Relatório", "Nenhum funcionário cadastrado.")
    else:
        messagebox.showerror("Erro", "Arquivo não encontrado.")

    tk.Button(frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=tela_inicial).pack(side="right", padx=10, pady=10)
def relatorio_admissao():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Relatório Geral de Admitidos", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de busca
    filtro_var = tk.StringVar()
    filtro_entry = tk.Entry(frame, textvariable=filtro_var, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    filtro_entry.pack(pady=10, ipady=5)
    
    
    def aplicar_filtro():
        query = filtro_var.get().lower()
        tree.delete(*tree.get_children())
        for _, row in df.iterrows():
            if any(query in str(row[col]).lower() for col in tree['columns']):
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
    
    tk.Button(frame, text="Filtrar", font=("Arial", 12, "bold"), width=20, height=2, bg="#008F68", fg="white", relief="ridge", bd=3, command=aplicar_filtro).pack(pady=10)

    
    # Criando um Frame para conter a Treeview e a Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    # Adicionando Scrollbars horizontal e vertical
    tree_scroll_y = tk.Scrollbar(tree_frame, orient="vertical")
    tree_scroll_y.pack(side="right", fill="y")

    tree_scroll_x = tk.Scrollbar(tree_frame, orient="horizontal")
    tree_scroll_x.pack(side="bottom", fill="x")

    colunas = ["Funcionário","Criar E-mail","Criar AD","Criar SISBR","Criar SIPAG","Criar ASSINATURA","Ramal WEBEX","Portal CCS A/O","CX compartilhada/lista dpt","Colocar lista3246.todos","Acessos Sisbr para central","Grupos outlook","Termo notebook","Cadastrar biometria"]
    
       
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=15, 
                        yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=120, anchor="center")

    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    tree.pack(expand=True, fill="both")
   
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    global df
    if os.path.exists(arquivo):
        df = pd.read_excel(arquivo, sheet_name="Admissao", dtype=str).fillna("")
        
        if not df.empty:
            for _, row in df.iterrows():
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
        else:
            messagebox.showinfo("Relatório", "Nenhum funcionário cadastrado.")
    else:
        messagebox.showerror("Erro", "Arquivo não encontrado.")

    tk.Button(frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=tela_inicial).pack(side="right", padx=10, pady=10)
    tk.Button(frame, text="Resolver Pendência", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=consultar_usuario_admissao).pack(side="right", padx=10, pady=10)
def relatorio_desligamento():
    limpar_tela()
    
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    label_titulo = tk.Label(frame, text="Relatório Geral de Desligados", font=("Arial", 18, "bold"), bg="#026440", fg="white", padx=10, pady=10)
    label_titulo.pack(fill="x", pady=10)

    # Campo de busca
    filtro_var = tk.StringVar()
    filtro_entry = tk.Entry(frame, textvariable=filtro_var, width=50, font=("Arial", 12), bd=3, relief="ridge", bg="#E8F5E9", fg="#003D1F")
    filtro_entry.pack(pady=10, ipady=5)
    
    
    def aplicar_filtro():
        query = filtro_var.get().lower()
        tree.delete(*tree.get_children())
        for _, row in df.iterrows():
            if any(query in str(row[col]).lower() for col in tree['columns']):
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
    
    tk.Button(frame, text="Filtrar", font=("Arial", 12, "bold"), width=20, height=2, bg="#008F68", fg="white", relief="ridge", bd=3, command=aplicar_filtro).pack(pady=10)

    
    # Criando um Frame para conter a Treeview e a Scrollbar
    tree_frame = tk.Frame(frame, bg="#003D1F")
    tree_frame.pack(pady=10, fill="both", expand=True)

    # Adicionando Scrollbars horizontal e vertical
    tree_scroll_y = tk.Scrollbar(tree_frame, orient="vertical")
    tree_scroll_y.pack(side="right", fill="y")

    tree_scroll_x = tk.Scrollbar(tree_frame, orient="horizontal")
    tree_scroll_x.pack(side="bottom", fill="x")

    colunas = ["Funcionário","Inativar SISBR", "Bloquear Sipag","Excluir Consórcio","Bloquear Prognum","Redirecionar e-mail","Backup e-mail", "Remover grupos e-mail", "Remover sharepoint", "Bloquear e-mail",
               "Excluir operador topdesk", "Inativar e mover AD","Inativar biometria","Desativar VPN","Comunicar CCS","Observações","Responsável"]
    
    tree = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=15, 
                        yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

    for coluna in colunas:
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=120, anchor="center")

    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    tree.pack(expand=True, fill="both")
   
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    global df
    if os.path.exists(arquivo):
        df = pd.read_excel(arquivo, sheet_name="Desligados", dtype=str).fillna("")
        
        if not df.empty:
            for _, row in df.iterrows():
                tree.insert("", "end", values=[row[col] for col in tree['columns']])
        else:
            messagebox.showinfo("Relatório", "Nenhum funcionário cadastrado.")
    else:
        messagebox.showerror("Erro", "Arquivo não encontrado.")

    tk.Button(frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=tela_inicial).pack(side="right", padx=10, pady=10)
    tk.Button(frame, text="Resolver Pendência", font=("Arial", 12, "bold"), width=20, height=2, bg="#00997B", fg="white", relief="ridge", bd=3, command=consultar_usuario_desligado).pack(side="right", padx=10, pady=10)

# Caminho do arquivo Excel
CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
ABA_ORIGEM = "Ativos"
ABA_DESTINO = "Inativos"
ABA_CHECKLIST_ADM = "Admissao"
ABA_CHECKLIST = "Desligados"

# Função para salvar checklist
def salvar_checklist(nome):
    responsavel = responsavel_var.get() .strip()    
    observacoes = text_observacoes.get("1.0", "end").strip()
    if responsavel.strip() == "" or responsavel.strip().lower() == "selecione...":
        messagebox.showwarning("Aviso", "Selecione o responsável pelo checklist.")
        return
    # Cria dicionário com o nome do funcionário e os status das tarefas
    linha = {"Funcionário": nome}
    for item in checklist_items:
        linha[item] = "Concluido" if checks[item].get() else "Pendente"
    linha["Observações"] = observacoes
    linha["Responsável"] = responsavel
        
    df_novo = pd.DataFrame([linha])  # DataFrame com uma linha
    if os.path.exists(CAMINHO_ARQUIVO):
        try:
            # Lê o arquivo existente e adiciona a nova linha
            with pd.ExcelWriter(CAMINHO_ARQUIVO, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                workbook = writer.book
                if ABA_CHECKLIST in workbook.sheetnames:
                    sheet = workbook[ABA_CHECKLIST]
                    existing_data = pd.read_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST)
                    df_final = pd.concat([existing_data, df_novo], ignore_index=True)
                    workbook.remove(sheet)
                    df_final.to_excel(writer, sheet_name=ABA_CHECKLIST, index=False)
                else:
                    df_novo.to_excel(writer, sheet_name=ABA_CHECKLIST, index=False)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar checklist: {str(e)}")
    else:
        df_novo.to_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST, index=False)

# Função que move o funcionário da aba Funcionarios para Inativos
def mover_para_inativos(nome_funcionario):
    if not os.path.exists(CAMINHO_ARQUIVO):
        messagebox.showerror("Erro", f"Arquivo '{CAMINHO_ARQUIVO}' não encontrado.")
        return

    wb = load_workbook(CAMINHO_ARQUIVO)
    if ABA_ORIGEM not in wb.sheetnames:
        messagebox.showerror("Erro", f"Aba '{ABA_ORIGEM}' não encontrada.")
        return

    ws_origem = wb[ABA_ORIGEM]
    dados = list(ws_origem.values)
    header = dados[0]
    linhas = dados[1:]

    linha_funcionario = None
    for i, linha in enumerate(linhas, start=2):  # começa da linha 2
        if nome_funcionario.lower() in str(linha[1]).lower():  # Coluna 2 (índice 1) = Nome
            linha_funcionario = (i, linha)
            break

    if not linha_funcionario:
        messagebox.showwarning("Atenção", f"Funcionário '{nome_funcionario}' não encontrado na aba '{ABA_ORIGEM}'.")
        return

    # Criar aba Inativos se não existir
    if ABA_DESTINO not in wb.sheetnames:
        ws_destino = wb.create_sheet(ABA_DESTINO)
        ws_destino.append(header)
    else:
        ws_destino = wb[ABA_DESTINO]

    # Copia e remove
    ws_destino.append(linha_funcionario[1])
    ws_origem.delete_rows(linha_funcionario[0], 1)

    wb.save(CAMINHO_ARQUIVO)
    messagebox.showinfo("Sucesso", f"Funcionário '{nome_funcionario}' movido para a aba'Inativos'.")
    tela_inicial()


# Função para salvar checklist
def salvar_checklist_a(observacoesa):
    #responsavel = responsavel_var.get() .strip()    
    observacoesa = text_observacoesa.get("1.0", "end").strip()
    #if responsavel.strip() == "" or responsavel.strip().lower() == "selecione...":
     #   messagebox.showwarning("Aviso", "Selecione o responsável pelo checklist.")
      #  return
    # Cria dicionário com o nome do funcionário e os status das tarefas
    linha = {"Funcionário": observacoesa}
    for item in checklist_items:
        linha[item] = "Concluido" if checks[item].get() else "Pendente"
    #linha["Nomess"] = observacoesa
    #linha["Responsável"] = responsavel
        
    df_novo = pd.DataFrame([linha])  # DataFrame com uma linha
    if os.path.exists(CAMINHO_ARQUIVO):
        try:
            # Lê o arquivo existente e adiciona a nova linha
            with pd.ExcelWriter(CAMINHO_ARQUIVO, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                workbook = writer.book
                if ABA_CHECKLIST_ADM in workbook.sheetnames:
                    sheet = workbook[ABA_CHECKLIST_ADM]
                    existing_data = pd.read_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST_ADM)
                    df_final = pd.concat([existing_data, df_novo], ignore_index=True)
                    workbook.remove(sheet)
                    df_final.to_excel(writer, sheet_name=ABA_CHECKLIST_ADM, index=False)
                else:
                    df_novo.to_excel(writer, sheet_name=ABA_CHECKLIST_ADM, index=False)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar checklist: {str(e)}")
    else:
        df_novo.to_excel(CAMINHO_ARQUIVO, sheet_name=ABA_CHECKLIST_ADM, index=False)
        
    messagebox.showinfo("Sucesso","Checklist salvo com sucesso!")
    tela_inicial()


# Função principal da tela de checklist
def checklist_desligamento():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Checklist de Desligamento", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")

    # Campo de pesquisa (nome, CPF, e-mail ou login)
    global entry_pesquisa
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa = tk.Entry(linha_pesquisa, font=("Arial", 12), width=42)
    entry_pesquisa.pack(side="left")
    entry_pesquisa.bind("<Return>", lambda event: buscar_funcionario_checklist())
          
    # Campo de exibição do nome encontrado (somente leitura)
    global entry_nome
    entry_nome = tk.Entry(frame, font=("Arial", 12), width=50, state="readonly")
    entry_nome.pack(pady=5)

    global checklist_items, checks
    checklist_items = [
        "Inativar SISBR",
        "Bloquear Sipag",
        "Excluir Consórcio",
        "Bloquear Prognum",
        "Redirecionar e-mail",
        "Backup e-mail",
        "Remover lista distribuição",
        "Remover grupos e-mail",
        "Remover sharepoint",
        "Bloquear e-mail",
        "Excluir operador topdesk",
        "Inativar e mover AD",
        "Inativar biometria",
        "Desativar VPN",
        "Comunicar CCS",        
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Tarefas", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)
    metade = len(checklist_items) // 2 + len(checklist_items) % 2  # Divide a lista

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Frame para responsável + observações lado a lado
    info_frame = tk.Frame(frame, bg="#003D1F")
    info_frame.pack(pady=10)

    # Responsável pelo checklist (label + combobox na mesma linha)
    global responsavel_var
    tk.Label(info_frame, text="Responsável:", font=("Arial", 14), bg="#003D1F", fg="white").grid(row=0, column=0, sticky="e", padx=(0,10))
    responsavel_var = tk.StringVar()
    combo_responsavel = ttk.Combobox(info_frame, textvariable=responsavel_var, values=["Denilson Oliveira", "Paulo Roberto", " Samuel Molina", "Reinaldo Camilo"], font=("Arial", 12), state="readonly", width=30)
    combo_responsavel.grid(row=0, column=1, sticky="w")
    combo_responsavel.set("Selecione...")

    # Observações (label + text box na mesma linha)
    global text_observacoes
    tk.Label(info_frame, text="Observações:", font=("Arial", 14), bg="#003D1F", fg="white").grid(row=1, column=0, sticky="ne", padx=(0,10), pady=(10, 0))
    text_observacoes = tk.Text(info_frame, width=55, height=1, font=("Arial", 12), wrap="word", bd=3, relief="ridge")
    text_observacoes.grid(row=1, column=1, sticky="w", pady=(10, 0))
        
    def ao_clicar_salvar():
        nome = entry_nome.get().strip()
        if not nome:
            messagebox.showwarning("Atenção", "Por favor, digite o nome do funcionário.")
            return
        salvar_checklist(nome)
        mover_para_inativos(nome)
    # Frame para botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)
    
    botoes = [
       # ("Buscar", buscar, "#008F68"),
        ("Salvar", ao_clicar_salvar, "#00B386"),
        ("Relatório/Pendências", relatorio_desligamento, "#00997B"),
        ("Voltar", tela_inicial, "#00997B")
    ]
    
    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)


def salvar_dados(entries):
    dados = {campo: entry.get() for campo, entry in entries.items()}
    
    if any(not valor for valor in dados.values()):
        messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
        return
    
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    salvar_em_excel(dados, arquivo)
    messagebox.showinfo("Sucesso", "Funcionário cadastrado com sucesso!")
    tela_inicial()

def salvar_em_excel(dados, arquivo):
    if os.path.exists(arquivo):
        df_existente = pd.read_excel(arquivo, dtype=str)
        df = pd.concat([df_existente, pd.DataFrame([dados])], ignore_index=True)
    else:
        df = pd.DataFrame([dados])
    
    df.to_excel(arquivo, index=False)
    
    wb = load_workbook(arquivo)
    ws = wb.active
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    fill1 = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill1 if i % 2 == 0 else fill2
        for cell in row:
            cell.fill = fill
    
    wb.save(arquivo)
def buscar_funcionario_checklist():
    termo_busca = entry_pesquisa.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite um nome, CPF, e-mail ou login para buscar.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_funcionarios = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Ativos")

        # Faz busca em várias colunas
        encontrados = df_funcionarios[
            df_funcionarios.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Erro", f"Nenhum funcionário encontrado com o termo '{termo_busca}'.")
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Nome"].tolist())
            messagebox.showerror("Erro", f"Mais de um funcionário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
        else:
            funcionario = encontrados.iloc[0]
            nome_encontrado = funcionario["Nome"]

            # Exibe no campo readonly
            entry_nome.config(state="normal")
            entry_nome.delete(0, tk.END)
            entry_nome.insert(0, nome_encontrado)
            entry_nome.config(state="readonly")

            messagebox.showinfo("Sucesso", f"Funcionário encontrado: {nome_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos funcionários:\n{e}")
# Variáveis globais
df_inativos_global = None
usuario_index_inativo = None

def normalizar_texto(texto):
    """Remove acentos, espaços extras e transforma em minúsculas para facilitar a comparação."""
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto

def consultar_usuario_desligado():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Consulta de Usuários Desligados", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")

    # Campo de pesquisa
    global entry_pesquisa_inativo
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite nome:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa_inativo = tk.Entry(linha_pesquisa, font=("Arial", 12), width=42)
    entry_pesquisa_inativo.pack(side="left")
    entry_pesquisa_inativo.bind("<Return>", lambda event: buscar_usuario_inativo())

    # Campo para mostrar o nome encontrado
    global entry_nome_inativo
    entry_nome_inativo = tk.Entry(frame, font=("Arial", 12), width=50, state="readonly")
    entry_nome_inativo.pack(pady=10)

    # Checklists
    global checklist_items, checks
    checklist_items = [
        "Inativar SISBR",
        "Bloquear Sipag",
        "Excluir Consórcio",
        "Bloquear Prognum",
        "Redirecionar e-mail",
        "Backup e-mail",
        "Remover lista distribuição",
        "Remover grupos e-mail",
        "Remover sharepoint",
        "Bloquear e-mail",
        "Excluir operador topdesk",
        "Inativar e mover AD",
        "Inativar biometria",
        "Desativar VPN",
        "Comunicar CCS",
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Checklist do Usuário", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)

    metade = len(checklist_items) // 2 + len(checklist_items) % 2

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    tk.Button(botoes_frame, text="Buscar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=buscar_usuario_inativo).grid(row=0, column=0, padx=10)
    tk.Button(botoes_frame, text="Salvar Checklist", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=salvar_checklist_inativo).grid(row=0, column=1, padx=10)
    tk.Button(botoes_frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=tela_inicial).grid(row=0, column=2, padx=10)

def buscar_usuario_inativo():
    global df_inativos_global, usuario_index_inativo

    termo_busca = entry_pesquisa_inativo.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite o nome.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_inativos = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Desligados")
        df_inativos_global = df_inativos

        colunas_map = {normalizar_texto(col): col for col in df_inativos.columns}

        encontrados = df_inativos[
            df_inativos.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Não encontrado", f"Nenhum usuário inativo encontrado com o termo '{termo_busca}'.")
            return
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Nome"].tolist())
            messagebox.showerror("Múltiplos resultados", f"Mais de um usuário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
            return
        else:
            usuario = encontrados.iloc[0]
            usuario_index_inativo = encontrados.index[0]

            nome_encontrado = usuario["Funcionário"]
            entry_nome_inativo.config(state="normal")
            entry_nome_inativo.delete(0, tk.END)
            entry_nome_inativo.insert(0, nome_encontrado)
            entry_nome_inativo.config(state="readonly")

            # Atualizar os checkboxes
            for item in checklist_items:
                nome_normalizado = normalizar_texto(item)
                coluna_correspondente = colunas_map.get(nome_normalizado)

                if coluna_correspondente:
                    status = str(usuario.get(coluna_correspondente, "")).strip().lower()
                    checks[item].set(status == "concluido")
                else:
                    checks[item].set(False)

            messagebox.showinfo("Encontrado", f"Usuário desligado encontrado: {nome_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos desligados:\n{e}")

def salvar_checklist_inativo():
    try:
        CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

        if df_inativos_global is None or usuario_index_inativo is None:
            messagebox.showerror("Erro", "Nenhum usuário inativo selecionado para salvar.")
            return

        for item in checklist_items:
            status = "Concluido" if checks[item].get() else "Pendente"
            if item not in df_inativos_global.columns:
                df_inativos_global[item] = ""  # Cria a coluna se não existir
            df_inativos_global.at[usuario_index_inativo, item] = status

        # Salva de volta somente a aba "Inativos"
        with pd.ExcelWriter(CAMINHO_ARQUIVO, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            df_inativos_global.to_excel(writer, sheet_name="Desligados", index=False)

        messagebox.showinfo("Sucesso", "Checklist salvo com sucesso!")
        tela_inicial()

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar o checklist:\n{e}")

def checklist_admissao():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Checklist de Admissão", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")
    info_frame = tk.Frame(frame, bg="#003D1F")
    info_frame.pack(pady=10)
    global text_observacoesa
    tk.Label(info_frame, text="Digite o nome completo:", font=("Arial", 14), bg="#003D1F", fg="white").grid(row=0, column=0, sticky="we", padx=(0,10), pady=(10, 0))
    text_observacoesa = tk.Text(info_frame, width=55, height=1, font=("Arial", 12), wrap="word", bd=3, relief="ridge")
    text_observacoesa.grid(row=1, column=0, sticky="w", pady=(5, 0))
     

    global checklist_items, checks
    checklist_items = [
        "Criar E-mail",
        "Criar AD",
        "Criar SISBR",
        "Criar SIPAG",
        "Criar ASSINATURA",
        "Ramal WEBEX",
        "Portal CCS A/O",
        "CX compartilhada/lista dpt",
        "Colocar lista3246.todos",
        "Acessos Sisbr para central",
        "Grupos outlook",
        "Termo notebook",
        "Cadastrar biometria",        
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Tarefas", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)
    metade = len(checklist_items) // 2 + len(checklist_items) % 2  # Divide a lista

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Frame para responsável + Nomess lado a lado
    info_frame = tk.Frame(frame, bg="#003D1F")
    info_frame.pack(pady=10)
    
       
    def ao_clicar_salvar():
        observacoesa = text_observacoesa.get("1.0", "end").strip()  # <-- Captura o texto digitado
        if not observacoesa:
            messagebox.showwarning("Atenção", "Por favor, digite o nome do funcionário.")
            return
        salvar_checklist_a(observacoesa)
    # Frame para botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)
    
    botoes = [
       # ("Buscar", buscar, "#008F68"),
        ("Salvar", ao_clicar_salvar, "#00B386"),
        ("Relatório/Pendências", relatorio_admissao, "#00997B"),
        ("Voltar", tela_inicial, "#00997B")
    ]
    
    for i, (text, command, color) in enumerate(botoes):
        tk.Button(botoes_frame, text=text, font=("Arial", 12, "bold"), width=20, height=2, bg=color, fg="white", relief="ridge", bd=3, command=command).grid(row=0, column=i, padx=10)


def salvar_dados(entries):
    dados = {campo: entry.get() for campo, entry in entries.items()}
    
    if any(not valor for valor in dados.values()):
        messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
        return
    
    arquivo = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")
    salvar_em_excel(dados, arquivo)
    messagebox.showinfo("Sucesso", "Funcionário cadastrado com sucesso!")
    tela_inicial()

def salvar_em_excel(dados, arquivo):
    if os.path.exists(arquivo):
        df_existente = pd.read_excel(arquivo, dtype=str)
        df = pd.concat([df_existente, pd.DataFrame([dados])], ignore_index=True)
    else:
        df = pd.DataFrame([dados])
    
    df.to_excel(arquivo, index=False)
    
    wb = load_workbook(arquivo)
    ws = wb.active
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    fill1 = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill1 if i % 2 == 0 else fill2
        for cell in row:
            cell.fill = fill
    
    wb.save(arquivo)
# Variáveis globais
df_admissaos_global = None
usuario_index_admissao = None

def normalizar_texto(texto):
    """Remove acentos, espaços extras e transforma em minúsculas para facilitar a comparação."""
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto

def consultar_usuario_admissao():
    limpar_tela()
    frame = tk.Frame(root, padx=20, pady=20, bg="#003D1F")
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Consulta de Usuários Admitidos", font=("Arial", 18, "bold"), bg="#026440", fg="white", pady=10).pack(fill="x")

    # Campo de pesquisa
    global entry_pesquisa_admissao
    linha_pesquisa = tk.Frame(frame, bg="#003D1F")
    linha_pesquisa.pack(pady=(10, 0))

    tk.Label(linha_pesquisa, text="Digite nome:", font=("Arial", 14, "bold"), bg="#003D1F", fg="white").pack(side="left", padx=(0, 10))

    entry_pesquisa_admissao = tk.Entry(linha_pesquisa, font=("Arial", 12), width=42)
    entry_pesquisa_admissao.pack(side="left")
    entry_pesquisa_admissao.bind("<Return>", lambda event: buscar_usuario_admissao())

    # Campo para mostrar o nome encontrado
    global entry_nome_admissao
    entry_nome_admissao = tk.Entry(frame, font=("Arial", 12), width=50, state="readonly")
    entry_nome_admissao.pack(pady=10)

    # Checklists
    global checklist_items, checks
    checklist_items = [
        "Criar E-mail",
        "Criar AD",
        "Criar SISBR",
        "Criar SIPAG",
        "Criar ASSINATURA",
        "Ramal WEBEX",
        "Portal CCS A/O",
        "CX compartilhada/lista dpt",
        "Colocar lista3246.todos",
        "Acessos Sisbr para central",
        "Grupos outlook",
        "Termo notebook",
        "Cadastrar biometria",
    ]

    checks = {}
    checklist_frame = tk.LabelFrame(frame, text="Checklist do Usuário", font=("Arial", 14, "bold"), bg="#003D1F", fg="white", padx=50, pady=10)
    checklist_frame.pack(pady=20)

    metade = len(checklist_items) // 2 + len(checklist_items) % 2

    for i, item in enumerate(checklist_items):
        col = 0 if i < metade else 1
        row = i if i < metade else i - metade

        checks[item] = tk.BooleanVar()
        chk = tk.Checkbutton(
            checklist_frame,
            text=item,
            variable=checks[item],
            font=("Arial", 12),
            bg="#003D1F",
            fg="white",
            selectcolor="#026440",
            activebackground="#003D1F",
            anchor="w"
        )
        chk.grid(row=row, column=col, sticky="w", padx=20, pady=5)

    # Botões
    botoes_frame = tk.Frame(frame, bg="#003D1F")
    botoes_frame.pack(pady=20)

    #tk.Button(botoes_frame, text="Buscar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=buscar_usuario_admissao).grid(row=0, column=0, padx=10)
    tk.Button(botoes_frame, text="Salvar Checklist", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=salvar_checklist_a_admissao).grid(row=0, column=1, padx=10)
    tk.Button(botoes_frame, text="Voltar", font=("Arial", 12, "bold"), width=20, height=2, bg="#007744", fg="white", relief="ridge", bd=3, command=tela_inicial).grid(row=0, column=2, padx=10)

def buscar_usuario_admissao():
    global df_admissaos_global, usuario_index_admissao

    termo_busca = entry_pesquisa_admissao.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite o nome.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_admissaos = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Admissao")
        df_admissaos_global = df_admissaos

        colunas_map = {normalizar_texto(col): col for col in df_admissaos.columns}

        encontrados = df_admissaos[
            df_admissaos.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Não encontrado", f"Nenhum usuário admitido  encontrado com o termo '{termo_busca}'.")
            return
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Nome"].tolist())
            messagebox.showerror("Múltiplos resultados", f"Mais de um usuário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
            return
        else:
            usuario = encontrados.iloc[0]
            usuario_index_admissao = encontrados.index[0]

            nome_encontrado = usuario["Funcionário"]
            entry_nome_admissao.config(state="normal")
            entry_nome_admissao.delete(0, tk.END)
            entry_nome_admissao.insert(0, nome_encontrado)
            entry_nome_admissao.config(state="readonly")

            # Atualizar os checkboxes
            for item in checklist_items:
                nome_normalizado = normalizar_texto(item)
                coluna_correspondente = colunas_map.get(nome_normalizado)

                if coluna_correspondente:
                    status = str(usuario.get(coluna_correspondente, "")).strip().lower()
                    checks[item].set(status == "concluido")
                else:
                    checks[item].set(False)

            messagebox.showinfo("Encontrado", f"Usuário admitido encontrado: {nome_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos admitidos:\n{e}")
    
def salvar_checklist_a_admissao():
    try:
        CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

        if df_admissaos_global is None or usuario_index_admissao is None:
            messagebox.showerror("Erro", "Nenhum usuário admitido selecionado para salvar.")
            return

        for item in checklist_items:
            status = "Concluido" if checks[item].get() else "Pendente"
            if item not in df_admissaos_global.columns:
                df_admissaos_global[item] = ""  # Cria a coluna se não existir
            df_admissaos_global.at[usuario_index_admissao, item] = status

        # Salva de volta somente a aba "Inativos"
        with pd.ExcelWriter(CAMINHO_ARQUIVO, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            df_admissaos_global.to_excel(writer, sheet_name="Admissao", index=False)

        messagebox.showinfo("Sucesso", "Checklist salvo com sucesso!")
        tela_inicial()

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar o checklist:\n{e}")
    
def buscar_funcionario_grupo():
    termo_busca = entry_pesquisa.get().strip()

    if not termo_busca:
        messagebox.showwarning("Aviso", "Digite um nome, CPF, e-mail ou login para buscar.")
        return

    CAMINHO_ARQUIVO = os.path.join("Y:\\Tecnologia", "Programa gestão tecnologia", "funcionarios.xlsx")

    try:
        df_funcionarios = pd.read_excel(CAMINHO_ARQUIVO, sheet_name="Ativos")

        # Faz busca em várias colunas
        encontrados = df_funcionarios[
            df_funcionarios.apply(lambda row: termo_busca.lower() in str(row).lower(), axis=1)
        ]

        if len(encontrados) == 0:
            messagebox.showerror("Erro", f"Nenhum funcionário encontrado com o termo '{termo_busca}'.")
        elif len(encontrados) > 1:
            nomes = "\n- " + "\n- ".join(encontrados["Login"].tolist())
            messagebox.showerror("Erro", f"Mais de um funcionário encontrado com o termo '{termo_busca}':{nomes}\n\nRefine a busca.")
        else:
            funcionario = encontrados.iloc[0]
            login_encontrado = funcionario["Login"]

            # Exibe no campo readonly
            entry_login.config(state="normal")
            entry_login.delete(0, tk.END)
            entry_login.insert(0, login_encontrado)
            entry_login.config(state="readonly")

            messagebox.showinfo("Sucesso", f"Funcionário encontrado: {login_encontrado}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao acessar os dados dos funcionários:\n{e}")


# Configuração inicial da janela
root = tk.Tk()
root.title("Gestão de Funcionários")
root.geometry("1100x800")

tela_inicial()

root.mainloop()
>>>>>>> 3233547d8f98473298e4433be7566257873820f7
>>>>>>> e8a803335aee9a806ce299f6e50df1f1541b199d
